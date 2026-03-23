#!/usr/bin/env python3
"""
Merge Inconsistent Summaries
================================

Reads Phase III Inconsistent IDs and Inconsistent Names Summary CSVs from a
directory, groups rows by a configurable key column per file, and merges each
group into a single row.

Merge behaviour:
- Where all rows in a group share the same value for a column, the merged cell
  contains that value once.
- Where rows differ, each unique value is stacked on a new line within the cell
  (duplicate values are suppressed; order of first appearance is preserved).
- Empty / NaN values are always suppressed when stacking.

Output is one Excel (.xlsx) file per processed CSV, written alongside the
input file and formatted with text wrapping, bold headers, and auto-sized
column widths.

Version 1.0 Changes:
- Initial implementation

Version 2.0 Changes:
- Extended to handle both Inconsistent IDs Summary and Inconsistent Names
  Summary files.
- Input changed from a single file path to an input directory; files are
  auto-discovered via configurable glob patterns.
- Output files written to the same directory as the inputs, with the .csv
  extension replaced by .xlsx.
- Console command renamed to merge-inconsistent-summaries.

Usage:
    # Using a YAML config file
    merge-inconsistent-summaries --config config/local/replay/merge_inconsistent_ids.yaml

    # Using CLI arguments directly
    merge-inconsistent-summaries --input-dir path/to/phase_iii/output

    # Dry run (no files written)
    merge-inconsistent-summaries --input-dir path/to/phase_iii/output --dry-run
"""

import argparse
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core import create_logger


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class MergeConfig:
    """Configuration for the merge operation."""

    input_dir: Path
    ids_pattern: str = "Replay_*_Inconsistent_IDs_Summary_*.csv"
    names_pattern: str = "Replay_*_Inconsistent_Names_Summary_*.csv"
    ids_group_column: str = "Reported IDs"
    names_group_column: str = "Reported Name & DOB"
    separator: str = "\n"
    dry_run: bool = False
    verbose: bool = False
    log_level: str = "INFO"


@dataclass
class MergeStats:
    """Statistics produced during a merge run."""

    input_rows: int = 0
    output_rows: int = 0
    groups_merged: int = 0          # groups that had more than one row
    groups_single: int = 0          # groups that were already unique
    columns_stacked: int = 0        # individual cell values that were stacked


# ---------------------------------------------------------------------------
# Configuration loading
# ---------------------------------------------------------------------------

class MergeConfigManager:
    """Loads MergeConfig from a YAML file, with optional CLI overrides."""

    DEFAULT_IDS_PATTERN: str = "Replay_*_Inconsistent_IDs_Summary_*.csv"
    DEFAULT_NAMES_PATTERN: str = "Replay_*_Inconsistent_Names_Summary_*.csv"
    DEFAULT_IDS_GROUP_COLUMN: str = "Reported IDs"
    DEFAULT_NAMES_GROUP_COLUMN: str = "Reported Name & DOB"
    DEFAULT_SEPARATOR: str = "\n"

    def load(
        self,
        config_file: Optional[str],
        cli_input_dir: Optional[str],
        cli_dry_run: bool,
        cli_verbose: bool,
        cli_log_level: Optional[str],
    ) -> MergeConfig:
        """
        Resolve configuration from YAML and CLI overrides.

        Priority (highest to lowest): CLI args > YAML > defaults.

        Args:
            config_file: Path to YAML config file, or None.
            cli_input_dir: --input-dir CLI value.
            cli_dry_run: --dry-run flag.
            cli_verbose: --verbose flag.
            cli_log_level: --log-level CLI value.

        Returns:
            Populated MergeConfig instance.

        Raises:
            FileNotFoundError: If a config file path is specified but missing.
            ValueError: If required paths cannot be resolved.
        """
        yaml_data: Dict[str, Any] = {}

        if config_file:
            config_path = Path(config_file)
            if not config_path.exists():
                raise FileNotFoundError(f"Config file not found: {config_path}")
            with open(config_path, encoding="utf-8") as fh:
                yaml_data = yaml.safe_load(fh) or {}

        paths = yaml_data.get("paths", {})
        files = yaml_data.get("files", {})
        merge = yaml_data.get("merge", {})
        options = yaml_data.get("options", {})
        processor = yaml_data.get("processor", {})

        # Resolve each value: CLI > YAML > default
        input_dir_raw = cli_input_dir or paths.get("input_dir")

        if not input_dir_raw:
            raise ValueError(
                "input_dir is required. Provide --input-dir or set paths.input_dir in config."
            )

        return MergeConfig(
            input_dir=Path(input_dir_raw),
            ids_pattern=files.get("ids_pattern", self.DEFAULT_IDS_PATTERN),
            names_pattern=files.get("names_pattern", self.DEFAULT_NAMES_PATTERN),
            ids_group_column=merge.get("ids_group_column", self.DEFAULT_IDS_GROUP_COLUMN),
            names_group_column=merge.get("names_group_column", self.DEFAULT_NAMES_GROUP_COLUMN),
            separator=merge.get("separator", self.DEFAULT_SEPARATOR),
            dry_run=cli_dry_run or options.get("dry_run", False),
            verbose=cli_verbose or options.get("verbose", False),
            log_level=cli_log_level or processor.get("log_level", "INFO"),
        )


# ---------------------------------------------------------------------------
# Core merge logic
# ---------------------------------------------------------------------------

class RowMerger:
    """
    Merges groups of rows from a DataFrame into single rows.

    Within each group (keyed by group_column):
    - Unique values are collected from each column, preserving first-appearance order.
    - Duplicate or blank values are suppressed.
    - If only one unique value exists, it is used directly.
    - If multiple unique values exist, they are joined with the configured separator.
    """

    def __init__(self, group_column: str, separator: str = "\n") -> None:
        """
        Initialise the RowMerger.

        Args:
            group_column: Column name to group rows by.
            separator: String used to join stacked values (default: newline).
        """
        self.group_column = group_column
        self.separator = separator

    def _merge_group(
        self,
        group_df: pd.DataFrame,
        all_columns: List[str],
    ) -> Dict[str, Any]:
        """
        Merge a group of rows into a single dict.

        Args:
            group_df: Subset of the DataFrame sharing the same group_column value.
            all_columns: Ordered list of all column names.

        Returns:
            Dict mapping column name to merged cell value.
        """
        row: Dict[str, Any] = {}
        for col in all_columns:
            if col == self.group_column:
                # Always a single shared value by definition
                row[col] = group_df[col].iloc[0]
                continue

            # Collect non-empty string representations, preserving order
            seen: Dict[str, None] = {}
            for raw in group_df[col]:
                if pd.isna(raw):
                    continue
                value = str(raw).strip()
                if value:
                    seen[value] = None  # dict preserves insertion order (Python 3.7+)

            unique_values = list(seen.keys())
            if len(unique_values) == 0:
                row[col] = ""
            elif len(unique_values) == 1:
                row[col] = unique_values[0]
            else:
                row[col] = self.separator.join(unique_values)

        return row

    def process(self, df: pd.DataFrame) -> tuple[pd.DataFrame, MergeStats]:
        """
        Process the full DataFrame, merging groups sharing the same group_column value.

        Args:
            df: Raw input DataFrame.

        Returns:
            Tuple of (merged DataFrame, MergeStats).

        Raises:
            KeyError: If group_column is not present in the DataFrame.
        """
        if self.group_column not in df.columns:
            raise KeyError(
                f"Group column '{self.group_column}' not found in CSV. "
                f"Available columns: {list(df.columns)}"
            )

        stats = MergeStats(input_rows=len(df))
        columns: List[str] = list(df.columns)
        merged_rows: List[Dict[str, Any]] = []

        for key, group_df in df.groupby(self.group_column, sort=False, dropna=False):
            merged_row = self._merge_group(group_df, columns)
            merged_rows.append(merged_row)

            if len(group_df) > 1:
                stats.groups_merged += 1
                # Count columns where stacking occurred
                for col, value in merged_row.items():
                    if isinstance(value, str) and self.separator in value:
                        stats.columns_stacked += 1
            else:
                stats.groups_single += 1

        merged_df = pd.DataFrame(merged_rows, columns=columns)
        stats.output_rows = len(merged_df)
        return merged_df, stats


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

# Maximum column width (in Excel character units) to prevent excessively wide columns
_MAX_COL_WIDTH: int = 40
# RGB hex fill colour for the header row (light steel blue)
_HEADER_FILL_COLOUR: str = "D0E4F1"


class ExcelExporter:
    """
    Writes a merged DataFrame to an Excel (.xlsx) file using openpyxl.

    Formatting applied:
    - Header row: bold font, light blue background fill, top-aligned.
    - Data cells: wrap_text=True, vertical top alignment.
    - Column widths: auto-sized from content, capped at _MAX_COL_WIDTH characters.
    - Row 1 frozen so the header stays visible when scrolling.
    """

    def __init__(self, separator: str = "\n") -> None:
        """
        Args:
            separator: Value separator used in merged cells (used for width calculation).
        """
        self.separator = separator

    def _estimate_col_width(self, series: pd.Series, header: str) -> int:
        """
        Estimate a sensible column width based on the header and the widest value.

        For multi-line cells, only the longest individual segment is used.

        Args:
            series: Column data.
            header: Column header string.

        Returns:
            Width in Excel character units, capped at _MAX_COL_WIDTH.
        """
        max_len = len(header)
        for val in series:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val)
            # If the cell has stacked lines, take the longest line
            segment_max = max((len(seg) for seg in s.split(self.separator)), default=0)
            if segment_max > max_len:
                max_len = segment_max
        return min(max_len + 2, _MAX_COL_WIDTH)  # +2 for padding

    def export(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        Write the DataFrame to an Excel file.

        Args:
            df: Merged DataFrame to export.
            output_path: Destination .xlsx path (parent directory created if needed).
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color=_HEADER_FILL_COLOUR,
            end_color=_HEADER_FILL_COLOUR,
            fill_type="solid",
        )
        header_alignment = Alignment(wrap_text=True, vertical="top")
        cell_alignment = Alignment(wrap_text=True, vertical="top")

        columns = list(df.columns)

        # --- Write and format the header row ---
        for col_idx, header in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Write and format data rows ---
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                # Write empty string for NaN/None; preserve all other values as strings
                if value is None or (isinstance(value, float) and pd.isna(value)):
                    cell_value: Any = ""
                elif isinstance(value, str):
                    cell_value = value
                else:
                    cell_value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = cell_alignment

        # --- Set column widths ---
        for col_idx, col_name in enumerate(columns, start=1):
            width = self._estimate_col_width(df[col_name], col_name)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # --- Freeze the header row ---
        ws.freeze_panes = "A2"

        wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI interface
# ---------------------------------------------------------------------------

def create_argument_parser() -> argparse.ArgumentParser:
    """
    Create the argument parser for merge-inconsistent-summaries.

    Returns:
        Configured ArgumentParser instance.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Merge Phase III Inconsistent IDs and Inconsistent Names Summary rows. "
            "Auto-discovers both CSV files from the input directory and writes one "
            "merged Excel file per CSV alongside the input."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Using a YAML config file
  merge-inconsistent-summaries --config config/local/replay/merge_inconsistent_ids.yaml

  # Specifying the input directory directly
  merge-inconsistent-summaries --input-dir path/to/phase_iii/output

  # Dry run — no files written
  merge-inconsistent-summaries --input-dir path/to/phase_iii/output --dry-run

  # Verbose output
  merge-inconsistent-summaries --config config/local/replay/merge_inconsistent_ids.yaml --verbose
        """,
    )

    parser.add_argument(
        "--config",
        type=str,
        metavar="PATH",
        help="Path to YAML configuration file (default: config/local/replay/merge_inconsistent_ids.yaml).",
    )
    parser.add_argument(
        "--input-dir",
        type=str,
        metavar="PATH",
        dest="input_dir",
        help="Directory containing the Phase III Inconsistent IDs and Names Summary CSV files.",
    )
    parser.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default=None,
        help="Logging verbosity level (default: INFO).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and merge but do not write the output files.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print per-group details during processing.",
    )

    return parser


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """
    Main entry point for the merge-inconsistent-summaries console script.

    Raises:
        SystemExit: On configuration error or fatal processing failure.
    """
    parser = create_argument_parser()
    args = parser.parse_args()

    # ---- Resolve config path (fall back to default local config if not specified) ----
    config_file = args.config
    if not config_file:
        default_config = (
            Path(__file__).parent.parent.parent
            / "config" / "local" / "replay" / "merge_inconsistent_ids.yaml"
        )
        if default_config.exists():
            print(f"Loading default configuration from {default_config}...")
            config_file = str(default_config)
        else:
            print(
                f"Error: No configuration specified and default config not found: {default_config}",
                file=sys.stderr,
            )
            sys.exit(1)

    # ---- Load config ----
    config_manager = MergeConfigManager()
    try:
        config = config_manager.load(
            config_file=config_file,
            cli_input_dir=args.input_dir,
            cli_dry_run=args.dry_run,
            cli_verbose=args.verbose,
            cli_log_level=args.log_level,
        )
    except (FileNotFoundError, ValueError) as exc:
        print(f"Configuration error: {exc}", file=sys.stderr)
        sys.exit(1)

    # ---- Configure logging ----
    logger = create_logger(__name__, log_dir="logs", log_level=config.log_level)
    log_level = getattr(logging, config.log_level.upper(), logging.INFO)
    logging.getLogger().setLevel(log_level)

    logger.info("merge-inconsistent-summaries starting")
    logger.info(f"Input directory: {config.input_dir}")

    if config.dry_run:
        logger.info("Dry run mode — no output files will be written.")

    # ---- Validate input directory ----
    if not config.input_dir.exists():
        logger.error(f"Input directory not found: {config.input_dir}")
        sys.exit(1)

    # ---- Discover input files ----
    ids_matches = sorted(config.input_dir.glob(config.ids_pattern))
    names_matches = sorted(config.input_dir.glob(config.names_pattern))

    if not ids_matches and not names_matches:
        logger.error(
            f"No matching files found in {config.input_dir}.\n"
            f"  IDs pattern:   {config.ids_pattern}\n"
            f"  Names pattern: {config.names_pattern}"
        )
        sys.exit(1)

    tasks: List[tuple[Path, str, str]] = []

    if ids_matches:
        if len(ids_matches) > 1:
            logger.warning(
                f"Multiple files match the IDs pattern '{config.ids_pattern}' — skipping:\n"
                + "\n".join(f"  {p}" for p in ids_matches)
            )
        else:
            tasks.append((ids_matches[0], config.ids_group_column, "IDs"))
    else:
        logger.warning(
            f"No file found matching IDs pattern '{config.ids_pattern}' in {config.input_dir} — skipping."
        )

    if names_matches:
        if len(names_matches) > 1:
            logger.warning(
                f"Multiple files match the Names pattern '{config.names_pattern}' — skipping:\n"
                + "\n".join(f"  {p}" for p in names_matches)
            )
        else:
            tasks.append((names_matches[0], config.names_group_column, "Names"))
    else:
        logger.warning(
            f"No file found matching Names pattern '{config.names_pattern}' in {config.input_dir} — skipping."
        )

    if not tasks:
        logger.error("No files could be processed.")
        sys.exit(1)

    exporter = ExcelExporter(separator=config.separator)

    for input_file, group_column, label in tasks:
        output_file = input_file.with_suffix(".xlsx")
        logger.info(f"[{label}] Input:        {input_file}")
        logger.info(f"[{label}] Output:       {output_file}")
        logger.info(f"[{label}] Group column: {group_column}")

        # ---- Read input CSV ----
        try:
            df = pd.read_csv(
                input_file,
                encoding="utf-8",
                dtype=str,              # Keep all values as strings; avoid type coercion
                keep_default_na=False,  # Treat empty cells as "" not NaN
            )
        except Exception as exc:  # noqa: BLE001
            logger.error(f"[{label}] Failed to read input CSV: {exc}")
            continue

        # Replace pandas NaN that might have crept in with empty string
        df = df.where(pd.notna(df), other="")

        logger.info(f"[{label}] Read {len(df):,} rows, {len(df.columns)} columns.")

        # ---- Merge ----
        merger = RowMerger(group_column=group_column, separator=config.separator)
        try:
            merged_df, stats = merger.process(df)
        except KeyError as exc:
            logger.error(f"[{label}] {exc}")
            continue

        if config.verbose:
            logger.debug(
                f"[{label}] Merge details: "
                f"{stats.groups_merged} groups merged (multiple rows), "
                f"{stats.groups_single} groups had a single row."
            )

        # ---- Dry run: just report ----
        if config.dry_run:
            logger.info(f"[{label}] Dry run — no file written.")
            _print_summary(stats, label, output_file, config, written=False)
            continue

        # ---- Export ----
        try:
            exporter.export(merged_df, output_file)
        except Exception as exc:  # noqa: BLE001
            logger.error(f"[{label}] Failed to write Excel file: {exc}")
            continue

        _print_summary(stats, label, output_file, config, written=True)

    logger.info("Done.")


def _print_summary(
    stats: MergeStats,
    label: str,
    output_file: Path,
    config: MergeConfig,
    *,
    written: bool,
) -> None:
    """
    Log a human-readable summary of the merge operation for one file.

    Args:
        stats: Statistics from the merge run.
        label: Short identifier for the file type (e.g. "IDs" or "Names").
        output_file: Path where the output was (or would have been) written.
        config: Resolved configuration.
        written: Whether the output file was actually written.
    """
    logger = create_logger(__name__, log_dir="logs", log_level=config.log_level)
    logger.info(
        f"\n"
        f"  [{label}] Merge Summary\n"
        f"  {'─' * 40}\n"
        f"  Input rows:          {stats.input_rows:>8,}\n"
        f"  Output rows:         {stats.output_rows:>8,}\n"
        f"  Rows removed:        {stats.input_rows - stats.output_rows:>8,}\n"
        f"  Groups merged (>1):  {stats.groups_merged:>8,}\n"
        f"  Single-row groups:   {stats.groups_single:>8,}\n"
        f"  Stacked cell values: {stats.columns_stacked:>8,}\n"
        f"  Output file:         {'(not written — dry run)' if not written else str(output_file)}"
    )


if __name__ == "__main__":
    main()
