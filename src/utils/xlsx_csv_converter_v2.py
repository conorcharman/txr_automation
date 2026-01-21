"""
XLSX to CSV Converter v2.0 - Enhanced with Recursive Processing

Converts XLSX files to CSV with multi-line cell splitting.
Supports recursive directory scanning with filters and smart path replacement.
Integrates with txr_replay_core for configuration and logging.

Author: Conor Charman
Date: 08 January 2026
Version: 2.0
"""

import csv
import argparse
import re
from pathlib import Path
from typing import Optional, Tuple, List, Any, Set
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Fallback to pandas if openpyxl not available
    import pandas as pd

from txr_replay_core import ConfigManager, ProcessingStats, create_logger


class XLSXConverterV2:
    """
    Enhanced XLSX to CSV converter with recursive scanning and filtering.
    """
    
    def __init__(
        self,
        parent_dir: Optional[Path] = None,
        input_dir: Optional[Path] = None,
        output_dir: Optional[Path] = None,
        logger = None,  # StructuredLogger type
        recursive: bool = False,
        filter_year: Optional[str] = None,
        filter_quarter: Optional[str] = None,
        filter_phases: Optional[List[str]] = None,
        dry_run: bool = False,
        force: bool = False
    ):
        """
        Initialize the enhanced XLSX converter.
        
        Args:
            parent_dir: Parent directory to scan recursively (xlsx→csv in same structure)
            input_dir: Single directory containing XLSX files (original mode)
            output_dir: Directory for CSV output files (used with input_dir)
            logger: StructuredLogger instance
            recursive: Enable recursive scanning
            filter_year: Filter by fiscal year (e.g., 'FY25')
            filter_quarter: Filter by quarter (e.g., 'Q3')
            filter_phases: List of phase names to include
            dry_run: Preview mode - don't actually convert files
            force: Overwrite existing CSV files without prompting
        """
        self.parent_dir = parent_dir
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.logger = logger
        self.recursive = recursive
        self.filter_year = filter_year
        self.filter_quarter = filter_quarter
        self.filter_phases = filter_phases or []
        self.dry_run = dry_run
        self.force = force
        self.stats = ProcessingStats()
        
        # Validate configuration
        if parent_dir and (input_dir or output_dir):
            raise ValueError("Cannot use both parent_dir and input_dir/output_dir modes")
        if not parent_dir and not input_dir:
            raise ValueError("Must specify either parent_dir or input_dir")
    
    @staticmethod
    def split_multiline_row(row: List[Any]) -> List[List[Any]]:
        """
        Split a single row containing multi-line cells into multiple rows.
        Optimized for streaming processing (processes one row at a time).
        
        Args:
            row: Single row to process
            
        Returns:
            List of rows (1 or more) after splitting multi-line cells
        """
        # Check if any cell contains newlines
        has_newlines = any(isinstance(val, str) and '\n' in val for val in row)
        
        if not has_newlines:
            return [row]
        
        # Split cells with newlines
        max_lines = 1
        split_cells = []
        
        for val in row:
            if isinstance(val, str) and '\n' in val:
                lines = val.split('\n')
                split_cells.append(lines)
                max_lines = max(max_lines, len(lines))
            else:
                split_cells.append([val])
        
        # Create new rows
        new_rows = []
        for line_idx in range(max_lines):
            new_row = []
            for cell_lines in split_cells:
                if line_idx < len(cell_lines):
                    new_row.append(cell_lines[line_idx])
                else:
                    new_row.append(cell_lines[-1])
            new_rows.append(new_row)
        
        return new_rows
    
    def matches_filters(self, path: Path) -> bool:
        """
        Check if a path matches the configured filters.
        
        Args:
            path: Path to check
            
        Returns:
            True if path matches all filters (or no filters set)
        """
        path_str = str(path)
        
        # Check fiscal year filter
        if self.filter_year:
            if self.filter_year not in path_str:
                return False
        
        # Check quarter filter
        if self.filter_quarter:
            if self.filter_quarter not in path_str:
                return False
        
        # Check phase filter
        if self.filter_phases:
            matched = False
            for phase in self.filter_phases:
                if phase in path_str:
                    matched = True
                    break
            if not matched:
                return False
        
        return True
    
    def get_output_path(self, xlsx_file: Path) -> Path:
        """
        Calculate the output CSV path based on the input path.
        For parent_dir mode: replaces 'xlsx' with 'csv' in the path
        For input_dir mode: uses output_dir directly
        
        Args:
            xlsx_file: Input XLSX file path
            
        Returns:
            Output CSV file path
        """
        if self.parent_dir:
            # Smart path replacement: xlsx → csv
            # Find the 'xlsx' folder in the path and replace with 'csv'
            parts = xlsx_file.parts
            new_parts = []
            
            for part in parts:
                if part.lower() == 'xlsx':
                    new_parts.append('csv')
                else:
                    new_parts.append(part)
            
            # Reconstruct path and change extension
            csv_path = Path(*new_parts).with_suffix('.csv')
            return csv_path
        else:
            # Original mode: output to specified output_dir
            if not self.output_dir:
                raise ValueError("output_dir must be specified in single directory mode")
            csv_filename = xlsx_file.stem + ".csv"
            return self.output_dir / csv_filename
    
    def find_xlsx_files(self) -> List[Path]:
        """
        Find all XLSX files based on configuration.
        
        Returns:
            List of XLSX file paths matching filters
        """
        xlsx_files = []
        
        if self.parent_dir:
            # Recursive mode with parent directory
            if self.recursive:
                pattern = "**/*.xlsx"
            else:
                pattern = "**/xlsx/*.xlsx"  # Only look in xlsx folders
            
            for xlsx_file in self.parent_dir.glob(pattern):
                # Apply filters
                if self.matches_filters(xlsx_file):
                    xlsx_files.append(xlsx_file)
        else:
            # Original single directory mode
            if not self.input_dir:
                return []
            if self.recursive:
                xlsx_files = list(self.input_dir.rglob("*.xlsx"))
            else:
                xlsx_files = list(self.input_dir.glob("*.xlsx"))
        
        return xlsx_files
    
    def convert_file_openpyxl(self, xlsx_file: Path, csv_file: Path) -> bool:
        """
        Convert a single XLSX file to CSV using openpyxl with streaming processing.
        Optimized for large files - processes rows one at a time instead of loading all into memory.
        
        Args:
            xlsx_file: Path to XLSX file
            csv_file: Path to output CSV file
            
        Returns:
            True if conversion successful, False otherwise
        """
        try:
            # Get file size for progress tracking
            file_size_mb = xlsx_file.stat().st_size / (1024 * 1024)
            
            base_dir = self.parent_dir if self.parent_dir else self.input_dir
            if base_dir:
                rel_path = xlsx_file.relative_to(base_dir)
                self.logger.info(f"Converting: {rel_path} ({file_size_mb:.1f} MB)")
            else:
                self.logger.info(f"Converting: {xlsx_file.name} ({file_size_mb:.1f} MB)")
            
            # Warn about large files
            if file_size_mb > 50:
                self.logger.warning(f"  ⚠ Large file detected - this may take several minutes")
            
            if self.dry_run:
                self.logger.info(f"  [DRY RUN] Would create: {csv_file}")
                self.stats.successful_matches += 1
                return True
            
            # Check if CSV already exists
            if csv_file.exists() and not self.force:
                self.logger.warning(f"  ⚠ CSV already exists (use --force to overwrite): {csv_file.name}")
                self.stats.errors += 1
                return False
            
            # Create output directory if needed
            csv_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Load workbook in read-only mode for better performance
            wb = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            
            # Open CSV file for streaming write
            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                row_count = 0
                output_row_count = 0
                progress_interval = 10000  # Log progress every 10k rows
                
                # Process rows in streaming fashion
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    row_count += 1
                    
                    if row_idx == 0:
                        # Write headers directly (no splitting)
                        writer.writerow(list(row))
                        output_row_count += 1
                    else:
                        # Process data row
                        processed_row = []
                        for cell_value in row:
                            # Format dates to DD/MM/YYYY
                            if isinstance(cell_value, datetime):
                                processed_row.append(cell_value.strftime('%d/%m/%Y'))
                            elif cell_value is None:
                                processed_row.append('')
                            else:
                                processed_row.append(str(cell_value) if not isinstance(cell_value, str) else cell_value)
                        
                        # Split multi-line cells and write immediately
                        split_rows = self.split_multiline_row(processed_row)
                        writer.writerows(split_rows)
                        output_row_count += len(split_rows)
                    
                    # Progress logging for large files
                    if row_count % progress_interval == 0:
                        self.logger.info(f"  Processing... {row_count:,} rows read, {output_row_count:,} rows written")
            
            # Close workbook to free memory
            wb.close()
            
            self.logger.info(f"  ✓ Created: {csv_file.name} ({row_count:,} input rows → {output_row_count:,} output rows)")
            self.stats.successful_matches += 1
            return True
            
        except Exception as e:
            self.logger.error(f"  ✗ Error converting {xlsx_file.name}: {str(e)}")
            self.stats.errors += 1
            return False
    
    def convert_file_pandas(self, xlsx_file: Path, csv_file: Path) -> bool:
        """
        Convert a single XLSX file to CSV using pandas (fallback method).
        
        Args:
            xlsx_file: Path to XLSX file
            csv_file: Path to output CSV file
            
        Returns:
            True if conversion successful, False otherwise
        """
        try:
            base_dir = self.parent_dir if self.parent_dir else self.input_dir
            if base_dir:
                rel_path = xlsx_file.relative_to(base_dir)
                self.logger.info(f"Converting: {rel_path}")
            else:
                self.logger.info(f"Converting: {xlsx_file.name}")
            
            if self.dry_run:
                self.logger.info(f"  [DRY RUN] Would create: {csv_file}")
                self.stats.successful_matches += 1
                return True
            
            # Check if CSV already exists
            if csv_file.exists() and not self.force:
                self.logger.warning(f"  ⚠ CSV already exists (use --force to overwrite): {csv_file.name}")
                self.stats.errors += 1
                return False
            
            # Read the XLSX file using pandas
            df = pd.read_excel(xlsx_file)
            
            # Format date columns to DD/MM/YYYY format (without time)
            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]' or pd.api.types.is_datetime64_any_dtype(df[col]):
                    # Convert datetime columns to DD/MM/YYYY format
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')
            
            # Convert dataframe to list of lists for split_multiline_rows
            headers = df.columns.tolist()
            data_rows = df.values.tolist()
            data_rows = self.split_multiline_rows(data_rows)
            
            # Create output directory if needed
            csv_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Write to CSV
            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(data_rows)
            
            self.logger.info(f"  ✓ Created: {csv_file.name}")
            self.stats.successful_matches += 1
            return True
            
        except Exception as e:
            self.logger.error(f"  ✗ Error converting {xlsx_file.name}: {str(e)}")
            self.stats.errors += 1
            return False
    
    def convert_all(self) -> Tuple[int, int]:
        """
        Convert all XLSX files based on configuration.
        
        Returns:
            Tuple of (successful_count, total_count)
        """
        # Find all XLSX files
        xlsx_files = self.find_xlsx_files()
        
        if not xlsx_files:
            self.logger.warning(f"No XLSX files found matching criteria")
            return 0, 0
        
        self.logger.info(f"Found {len(xlsx_files)} XLSX file(s) to convert")
        if self.dry_run:
            self.logger.info("[DRY RUN MODE - No files will be modified]")
        self.stats.processed_files = len(xlsx_files)
        
        # Choose conversion method based on availability
        if OPENPYXL_AVAILABLE:
            self.logger.info("Using openpyxl (optimized)")
            convert_method = self.convert_file_openpyxl
        else:
            self.logger.info("Using pandas (fallback)")
            convert_method = self.convert_file_pandas
        
        # Convert each file
        for xlsx_file in xlsx_files:
            csv_file = self.get_output_path(xlsx_file)
            convert_method(xlsx_file, csv_file)
        
        # Log summary statistics
        self.logger.log_stats(self.stats)
        
        return self.stats.successful_matches, self.stats.processed_files


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='XLSX to CSV Converter v2.0 - Enhanced with recursive processing',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Recursive mode with parent directory (xlsx → csv in same structure)
  replay-xlsx-converter-v2 --parent-dir C:\\Data\\txr_replay_automation --recursive
  
  # Filter by fiscal year and quarter
  replay-xlsx-converter-v2 --parent-dir C:\\Data\\txr_replay_automation --filter-year FY25 --filter-quarter Q3
  
  # Filter specific phases
  replay-xlsx-converter-v2 --parent-dir C:\\Data\\txr_replay_automation --filter-phase phase_ii phase_iii
  
  # Dry run to preview what would be converted
  replay-xlsx-converter-v2 --parent-dir C:\\Data\\txr_replay_automation --dry-run
  
  # Original single-directory mode
  replay-xlsx-converter-v2 --input-dir ./data/xlsx --output-dir ./data/csv
  
  # Use configuration file
  replay-xlsx-converter-v2 --config config/xlsx_converter_v2.yaml
        """
    )
    
    parser.add_argument(
        '--config',
        type=str,
        help='Path to configuration YAML file'
    )
    
    parser.add_argument(
        '--mode',
        type=int,
        choices=[1, 2],
        help='Conversion mode: 1=Recursive parent directory, 2=Single directory'
    )
    
    parser.add_argument(
        '--parent-dir',
        type=str,
        help='Parent directory to scan recursively (xlsx→csv in same structure)'
    )
    
    parser.add_argument(
        '--input-dir',
        type=str,
        help='Single directory containing XLSX files (original mode)'
    )
    
    parser.add_argument(
        '--output-dir',
        type=str,
        help='Directory for CSV output files (used with --input-dir)'
    )
    
    parser.add_argument(
        '--recursive',
        action='store_true',
        help='Enable recursive scanning of subdirectories'
    )
    
    parser.add_argument(
        '--filter-year',
        type=str,
        help='Filter by fiscal year (e.g., FY25)'
    )
    
    parser.add_argument(
        '--filter-quarter',
        type=str,
        help='Filter by quarter (e.g., Q3)'
    )
    
    parser.add_argument(
        '--filter-phase',
        nargs='+',
        help='Filter by phase names (e.g., phase_ii phase_iii reference)'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview mode - show what would be converted without actually converting'
    )
    
    parser.add_argument(
        '--force',
        action='store_true',
        help='Overwrite existing CSV files without prompting'
    )
    
    parser.add_argument(
        '--log-level',
        type=str,
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Logging level (default: INFO)'
    )
    
    return parser.parse_args()


def main():
    """Main execution function."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Load configuration
    config_dict = {}
    
    if args.config:
        # Load from YAML file
        config_dict = ConfigManager.load_from_yaml(args.config)
    else:
        # Try to load default local config
        default_config = Path(__file__).parent.parent.parent / "config" / "local" / "utils" / "xlsx_converter_v2.yaml"
        if default_config.exists():
            print(f"Loading default configuration from {default_config}...")
            config_dict = ConfigManager.load_from_yaml(str(default_config))
    
    # Override with command line arguments if provided
    if not config_dict.get('paths'):
        config_dict['paths'] = {}
    if not config_dict.get('processing'):
        config_dict['processing'] = {}
    if not config_dict.get('filters'):
        config_dict['filters'] = {}
    
    # Handle mode selection
    if args.mode:
        config_dict['mode'] = args.mode
    
    # Apply mode-based configuration
    selected_mode = config_dict.get('mode')
    if selected_mode == 1:
        # Mode 1: Use mode1_paths configuration
        if 'mode1_paths' in config_dict:
            config_dict['paths'] = config_dict['mode1_paths'].copy()
    elif selected_mode == 2:
        # Mode 2: Use mode2_paths configuration
        if 'mode2_paths' in config_dict:
            config_dict['paths'] = config_dict['mode2_paths'].copy()
    
    # Command line overrides (these take precedence over mode)
    if args.parent_dir:
        config_dict['paths']['parent_dir'] = args.parent_dir
        config_dict['paths'].pop('input_dir', None)
        config_dict['paths'].pop('output_dir', None)
    if args.input_dir:
        config_dict['paths']['input_dir'] = args.input_dir
        config_dict['paths'].pop('parent_dir', None)
    if args.output_dir:
        config_dict['paths']['output_dir'] = args.output_dir
    if args.recursive:
        config_dict['processing']['recursive'] = True
    if args.filter_year:
        config_dict['filters']['fiscal_year'] = args.filter_year
    if args.filter_quarter:
        config_dict['filters']['quarter'] = args.filter_quarter
    if args.filter_phase:
        config_dict['filters']['phases'] = args.filter_phase
    if args.dry_run:
        config_dict['processing']['dry_run'] = True
    if args.force:
        config_dict['processing']['force'] = True
    
    # Extract configuration
    parent_dir = config_dict.get('paths', {}).get('parent_dir')
    input_dir = config_dict.get('paths', {}).get('input_dir')
    output_dir = config_dict.get('paths', {}).get('output_dir')
    recursive = config_dict.get('processing', {}).get('recursive', False)
    filter_year = config_dict.get('filters', {}).get('fiscal_year')
    filter_quarter = config_dict.get('filters', {}).get('quarter')
    filter_phases = config_dict.get('filters', {}).get('phases')
    dry_run = config_dict.get('processing', {}).get('dry_run', False)
    force = config_dict.get('processing', {}).get('force', False)
    
    # Validate configuration
    if not parent_dir and not input_dir:
        print("Error: Must specify either --parent-dir or --input-dir")
        print("Use --help for usage information")
        return 1
    
    if parent_dir and (input_dir or output_dir):
        print("Error: Cannot use both --parent-dir and --input-dir/--output-dir")
        return 1
    
    if input_dir and not output_dir:
        print("Error: --output-dir required when using --input-dir")
        return 1
    
    # Convert to Path objects
    if parent_dir:
        parent_dir = Path(parent_dir)
        if not parent_dir.exists():
            print(f"Error: Parent directory does not exist: {parent_dir}")
            return 1
    
    if input_dir:
        input_dir = Path(input_dir)
        if not input_dir.exists():
            print(f"Error: Input directory does not exist: {input_dir}")
            return 1
        output_dir = Path(output_dir)
    
    # Setup logging
    log_output = config_dict.get('paths', {}).get('log_output', str(parent_dir or output_dir))
    logger = create_logger('xlsx_csv_converter_v2', log_output, args.log_level)
    
    # Log startup information
    logger.info("=" * 80)
    logger.info("XLSX to CSV Converter v2.0 - Enhanced Edition")
    logger.info("=" * 80)
    
    if parent_dir:
        logger.info(f"Parent directory: {parent_dir}")
        logger.info(f"Mode:             Recursive (xlsx→csv in same structure)")
    else:
        logger.info(f"Input directory:  {input_dir}")
        logger.info(f"Output directory: {output_dir}")
        logger.info(f"Mode:             Single directory")
    
    if filter_year:
        logger.info(f"Filter year:      {filter_year}")
    if filter_quarter:
        logger.info(f"Filter quarter:   {filter_quarter}")
    if filter_phases:
        logger.info(f"Filter phases:    {', '.join(filter_phases)}")
    if dry_run:
        logger.info(f"Dry run:          ENABLED (preview only)")
    if force:
        logger.info(f"Force overwrite:  ENABLED")
    
    logger.info(f"Log level:        {args.log_level}")
    
    # Create converter and run
    try:
        converter = XLSXConverterV2(
            parent_dir=parent_dir,
            input_dir=input_dir,
            output_dir=output_dir,
            logger=logger,
            recursive=recursive,
            filter_year=filter_year,
            filter_quarter=filter_quarter,
            filter_phases=filter_phases,
            dry_run=dry_run,
            force=force
        )
        
        successful, total = converter.convert_all()
        
        # Log final summary
        logger.info("=" * 80)
        logger.info("Conversion Complete")
        logger.info("=" * 80)
        logger.info(f"Successfully converted: {successful}/{total} files")
        
        if dry_run:
            logger.info("(Dry run - no files were actually modified)")
        
        return 0 if successful == total else 1
        
    except ValueError as e:
        print(f"Configuration error: {e}")
        return 1
    except Exception as e:
        print(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    exit(main())
