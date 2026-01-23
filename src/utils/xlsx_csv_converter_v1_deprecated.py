"""
XLSX to CSV Converter - Refactored v2.0

Converts XLSX files to CSV with multi-line cell splitting.
Integrates with txr_replay_core for configuration and logging.

Author: Conor Charman
Date: 24 December 2025
Version: 2.0
"""

import csv
import argparse
from pathlib import Path
from typing import Optional, Tuple, List, Any
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Fallback to pandas if openpyxl not available
    import pandas as pd

from txr_replay_core import ConfigManager, ProcessingStats, create_logger


class XLSXConverter:
    """
    XLSX to CSV converter with multi-line cell handling.
    """
    
    def __init__(self, input_dir: Path, output_dir: Path, logger):
        """
        Initialize the XLSX converter.
        
        Args:
            input_dir: Directory containing XLSX files
            output_dir: Directory for CSV output files
            logger: StructuredLogger instance
        """
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.logger = logger
        self.stats = ProcessingStats()
    
    @staticmethod
    def split_multiline_rows(rows: List[List[Any]]) -> List[List[Any]]:
        """
        Split rows containing multi-line cells (cells with newlines) into separate rows.
        For each row, if any cell contains newlines, create multiple rows where:
        - Cells with newlines are split across the new rows
        - Cells without newlines are copied to all new rows
        
        Args:
            rows: List of row lists to process
            
        Returns:
            List of rows with multi-line cells split into separate rows
        """
        new_rows = []
        
        for row in rows:
            # Check if any cell in this row contains newlines
            has_newlines = any(isinstance(val, str) and '\n' in val for val in row)
            
            if not has_newlines:
                # No newlines, keep row as is
                new_rows.append(row)
            else:
                # Split cells with newlines
                # First, determine the maximum number of lines in any cell
                max_lines = 1
                split_cells = []
                
                for val in row:
                    if isinstance(val, str) and '\n' in val:
                        # Split by newline
                        lines = val.split('\n')
                        split_cells.append(lines)
                        max_lines = max(max_lines, len(lines))
                    else:
                        # Single value, will be repeated
                        split_cells.append([val])
                
                # Create new rows
                for line_idx in range(max_lines):
                    new_row = []
                    for cell_lines in split_cells:
                        if line_idx < len(cell_lines):
                            # Use the corresponding line
                            new_row.append(cell_lines[line_idx])
                        else:
                            # Use the last line if we've run out
                            new_row.append(cell_lines[-1])
                    new_rows.append(new_row)
        
        return new_rows
    
    def convert_file_openpyxl(self, xlsx_file: Path) -> bool:
        """
        Convert a single XLSX file to CSV using openpyxl (faster, no pandas overhead).
        
        Args:
            xlsx_file: Path to XLSX file
            
        Returns:
            True if conversion successful, False otherwise
        """
        try:
            self.logger.info(f"Converting: {xlsx_file.name}")
            
            # Load workbook in read-only mode for better performance
            wb = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
            ws = wb.active
            
            # Read all rows into memory
            rows = []
            headers = None
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # Store headers
                    headers = list(row)
                    rows.append(headers)
                else:
                    # Process data rows
                    processed_row = []
                    for cell_value in row:
                        # Format dates to DD/MM/YYYY
                        if isinstance(cell_value, datetime):
                            processed_row.append(cell_value.strftime('%d/%m/%Y'))
                        elif cell_value is None:
                            processed_row.append('')
                        else:
                            processed_row.append(str(cell_value) if not isinstance(cell_value, str) else cell_value)
                    rows.append(processed_row)
            
            wb.close()
            
            # Split multi-line rows
            if len(rows) > 1:  # Has data beyond headers
                headers = rows[0]
                data_rows = rows[1:]
                data_rows = self.split_multiline_rows(data_rows)
                rows = [headers] + data_rows
            
            # Write to CSV
            csv_filename = xlsx_file.stem + ".csv"
            csv_filepath = self.output_dir / csv_filename
            
            with open(csv_filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            
            self.logger.info(f"  ✓ Successfully converted to: {csv_filename}")
            self.stats.successful_matches += 1
            return True
            
        except Exception as e:
            self.logger.logger.error(f"  ✗ Error converting {xlsx_file.name}: {str(e)}")
            self.stats.errors += 1
            return False
    
    def convert_file_pandas(self, xlsx_file: Path) -> bool:
        """
        Convert a single XLSX file to CSV using pandas (fallback method).
        
        Args:
            xlsx_file: Path to XLSX file
            
        Returns:
            True if conversion successful, False otherwise
        """
        try:
            self.logger.info(f"Converting: {xlsx_file.name}")
            
            # Read the XLSX file using pandas
            df = pd.read_excel(xlsx_file)
            
            # Format date columns to DD/MM/YYYY format (without time)
            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]' or pd.api.types.is_datetime64_any_dtype(df[col]):
                    # Convert datetime columns to DD/MM/YYYY format
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')
            
            # Convert dataframe to list of lists for split_multiline_rows
            headers = df.columns.tolist()
            data_rows = df.values.tolist()
            data_rows = self.split_multiline_rows(data_rows)
            
            # Create the output filename by changing the extension
            csv_filename = xlsx_file.stem + ".csv"
            csv_filepath = self.output_dir / csv_filename
            
            # Write to CSV
            with open(csv_filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(data_rows)
            
            self.logger.info(f"  ✓ Successfully converted to: {csv_filename}")
            self.stats.successful_matches += 1
            return True
            
        except Exception as e:
            self.logger.logger.error(f"  ✗ Error converting {xlsx_file.name}: {str(e)}")
            self.stats.errors += 1
            return False
    
    def convert_all(self) -> Tuple[int, int]:
        """
        Convert all XLSX files in input directory.
        
        Returns:
            Tuple of (successful_count, total_count)
        """
        # Find all XLSX files
        xlsx_files = list(self.input_dir.glob("*.xlsx"))
        
        if not xlsx_files:
            self.logger.logger.warning(f"No XLSX files found in {self.input_dir}")
            return 0, 0
        
        self.logger.info(f"Found {len(xlsx_files)} XLSX file(s) to convert")
        self.stats.processed_files = len(xlsx_files)
        
        # Choose conversion method based on availability
        if OPENPYXL_AVAILABLE:
            self.logger.info("Using openpyxl (optimized)")
            convert_method = self.convert_file_openpyxl
        else:
            self.logger.info("Using pandas (fallback)")
            convert_method = self.convert_file_pandas
        
        # Convert each file
        for xlsx_file in xlsx_files:
            convert_method(xlsx_file)
        
        # Log summary statistics
        self.logger.log_stats(self.stats)
        
        return self.stats.successful_matches, self.stats.processed_files


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='XLSX to CSV Converter with multi-line cell handling',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use configuration file
  python -m src.utils.xlsx_csv_converter --config config/xlsx_converter.yaml
  
  # Use command line arguments
  python -m src.utils.xlsx_csv_converter --input-dir ./data/xlsx --output-dir ./data/csv
  
  # Use environment variables
  export TXR_INPUT_DIR=./data/xlsx
  export TXR_OUTPUT_DIR=./data/csv
  python -m src.utils.xlsx_csv_converter --use-env
        """
    )
    
    parser.add_argument(
        '--config',
        type=str,
        help='Path to configuration YAML file'
    )
    
    parser.add_argument(
        '--use-env',
        action='store_true',
        help='Load configuration from environment variables (TXR_* prefix)'
    )
    
    parser.add_argument(
        '--input-dir',
        type=str,
        help='Directory containing XLSX files (overrides config)'
    )
    
    parser.add_argument(
        '--output-dir',
        type=str,
        help='Directory for CSV output files (overrides config)'
    )
    
    parser.add_argument(
        '--log-level',
        type=str,
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
        default='INFO',
        help='Logging level (default: INFO)'
    )
    
    return parser.parse_args()


def main():
    """Main execution function."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Load configuration
    config_dict = {}
    
    if args.config:
        # Load from YAML file
        config_dict = ConfigManager.load_from_yaml(args.config)
    elif args.use_env:
        # Load from environment variables
        config_dict = ConfigManager.load_from_env()
    else:
        # Default configuration path (use local config)
        default_config = Path(__file__).parent.parent.parent / "config" / "local" / "utils" / "xlsx_converter.yaml"
        if default_config.exists():
            print(f"Loading default configuration from {default_config}...")
            config_dict = ConfigManager.load_from_yaml(str(default_config))
        else:
            # Continue without config - will require command line args
            config_dict = {}
    
    # Override with command line arguments if provided
    if not config_dict.get('paths'):
        config_dict['paths'] = {}
    
    if args.input_dir:
        config_dict['paths']['input_dir'] = args.input_dir
    if args.output_dir:
        config_dict['paths']['output_dir'] = args.output_dir
    
    # Validate required paths
    if 'input_dir' not in config_dict.get('paths', {}):
        print("Error: input_dir must be specified via --input-dir, config file, or environment variable TXR_INPUT_DIR")
        return 1
    
    if 'output_dir' not in config_dict.get('paths', {}):
        print("Error: output_dir must be specified via --output-dir, config file, or environment variable TXR_OUTPUT_DIR")
        return 1
    
    # Extract paths
    input_dir = Path(config_dict['paths']['input_dir'])
    output_dir = Path(config_dict['paths']['output_dir'])
    
    # Check if input directory exists
    if not input_dir.exists():
        print(f"Error: Input directory does not exist: {input_dir}")
        return 1
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Setup logging
    log_output = config_dict.get('paths', {}).get('log_output', str(output_dir))
    logger = create_logger('xlsx_csv_converter', log_output, args.log_level)
    
    # Log startup information
    logger.info("=" * 80)
    logger.info("XLSX to CSV Converter v2.0")
    logger.info("=" * 80)
    logger.info(f"Input directory:  {input_dir}")
    logger.info(f"Output directory: {output_dir}")
    logger.info(f"Log level:        {args.log_level}")
    
    # Create converter and run
    converter = XLSXConverter(input_dir, output_dir, logger)
    successful, total = converter.convert_all()
    
    # Log final summary
    logger.info("=" * 80)
    logger.info("Conversion Complete")
    logger.info("=" * 80)
    logger.info(f"Successfully converted: {successful}/{total} files")
    
    return 0 if successful == total else 1


if __name__ == "__main__":
    exit(main())
    input("\nPress Enter to close...")
