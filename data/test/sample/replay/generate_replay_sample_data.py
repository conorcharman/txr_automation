#!/usr/bin/env python3
"""
Generate Replay Sample Data (pipeline-aligned version)
=======================================================

Creates sample files for all four replay scripts where transaction
references, person IDs, and corrections are drawn from the accuracy
testing sample data so the full pipeline can be exercised end-to-end.

Pipeline flow
-------------
SQL extract -> accuracy testing extract CSV
            -> accuracy testing template CSV  (corrections filled in)
            -> Phase 2 replay:
                KR Analysis XLSX  (one row per transaction in the extract)
                incident template  (corrections the processor will write back)
            -> Phase 3 replay:
                Inconsistent IDs Summary  (persons from 7_66 extract)
                Inconsistent Names Summary (persons from 16_20 extract)
            -> Phase 3 Final Lookup:
                processed summary CSVs + UnaVista CSV

Accuracy testing data used
--------------------------
7_39   buyer ID validation      TXN001-TXN019
16_23  seller ID validation     TXN101-TXN119
35_3   net amount               TXN400-TXN411
7_66   inconsistent buyer ID    TXN200-TXN217
16_20  inconsistent seller ID   TXN300-TXN317

Output
------
data/test/sample/replay/
+-phase_2_feedback/
|   +-1903a~G15~P2_0-7~7_39~KR Final Analysis_Data_1 OF 1.xlsx
|   +-1903a~G15~P2_30-39~35_3~KR Final Analysis_Data_1 OF 1.xlsx
|   +-1903a~G15~P2_0-19~7_66+16_20~KR Final Analysis_Data_1 OF 1.xlsx
+-phase_2_incident_templates/
|   +-FY26 Q1 7_39.csv
|   +-FY26 Q1 16_23.csv
|   +-FY26 Q1 35_3.csv
|   +-FY26 Q1 7_66.csv
|   +-FY26 Q1 16_20.csv
+-phase_3_feedback/
|   +-Replay_2025Q3_PHASE 3_Inconsistent_IDs_Summary_FINAL.csv
|   +-Replay_2025Q3_PHASE 3_Inconsistent_Names_Summary_FINAL.csv
+-phase_3_final_lookup/
    +-Replay_2025Q3_Inconsistent_IDs_Summary_FINAL.csv
    +-Replay_2025Q3_Inconsistent_Names_Summary_FINAL.csv
    +-UnaVista_MiFIR_Manual_Corrections_423_20180406111252.(264).csv

Usage
-----
    python data/test/sample/replay/generate_replay_sample_data.py
"""

import csv
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).parent

TOTAL_KR_COLS = 104
EXEC_ENTITY = "213800SAMPLE0001LEI1"
INTC = "INTC"


# ===========================================================================
# Helpers
# ===========================================================================

def _make_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _write_csv(filepath: Path, rows: list[list]) -> None:
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"  Created: {filepath.name}  ({len(rows) - 1} data rows)")


def _pad_row(row: list, total_cols: int) -> list:
    return row + [None] * (total_cols - len(row))


def _write_kr_xlsx(filepath: Path, header: list, data_rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "in1"
    fill = PatternFill("solid", fgColor="BDD7EE")
    ws.append(_pad_row(header, TOTAL_KR_COLS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in data_rows:
        ws.append(_pad_row(row, TOTAL_KR_COLS))
    ws_list = wb.create_sheet("List")
    ws_list.append(["List"])
    ws_list.append(["Yes"])
    ws_list.append(["No"])
    wb.save(filepath)
    print(f"  Created: {filepath.name}  ({len(data_rows)} data rows)")


# Standard incident template header (read by phase_2_processor via config column names)
_TMPL_HDR = [
    "Transaction Reference",
    "Correction",
    "Correction Field",
    "Agree With Correction",
    "Suggested Correction",
    "Suggested Correction Field",
    "Error",
]

# ===========================================================================
# KR Analysis XLSX column headers
# ===========================================================================

# Single-incident (AGREES=8, CORRECTION_FIELD=9, CORRECTION_VALUE=10, TXN_REF=13)
_KR_SINGLE_BASE = [
    "KR_Incident_Code",                           # 0
    "KR_Incident_Codes_Grouping",                 # 1
    "KR_IssueCode",                               # 2
    "KR_Diagnosis",                               # 3
    "KR_Diagnostics_key",                         # 4
    "KR_Proposed_Correction_Field",               # 5
    "KR_Proposed_Correction_Value",               # 6
    "KR_Revised_Reportability",                   # 7
    "Client_Agrees_With_KR_Proposed_Correction",  # 8  AGREES
    "Client_Final_Correction_Field",              # 9  CORRECTION_FIELD
    "Client_Final_Correction_Value",              # 10 CORRECTION_VALUE
    "KR_RECORD_KEY",                              # 11
    "Executing entity identification code",       # 12
    "Transaction reference number",               # 13 MATCH KEY
    "KR_REPORTABLE",                              # 14
    "KR_REPORTABLE_REASON",                       # 15
]

_KR_SINGLE_IDENTITY = _KR_SINGLE_BASE + [
    "Country of the branch for the buyer",   # 16
    "Buyer identification code",             # 17
    "Buyer_Name",                            # 18
    "Country of the branch for the seller",  # 19
    "Seller identification code",            # 20
    "Seller_Name",                           # 21
    "Seller decision maker code",            # 22
    "Trading capacity",                      # 23
    "Venue",                                 # 24
    "Type of buyer identification code",     # 25
    "Type of seller identification code",    # 26
    "Type of buyer decision maker code",     # 27
    "Type of seller decision maker code",    # 28
    "KR_Prior_Group",                        # 29
    "KR_Prior_Group_Phase",                  # 30
    "KR_Prior_Group_Correction",             # 31
]

_KR_SINGLE_NET_AMOUNT = _KR_SINGLE_BASE + [
    "KR_Reportable_Description",      # 16
    "KR_Instrument_Classification",   # 17
    "Instrument identification code", # 18
    "Net amount",                     # 19
    "Net_amount_delta",               # 20
    "Type of quantity",               # 21
    "Quantity currency",              # 22
    "Quantity_str",                   # 23
    "Type of price",                  # 24
    "Price_str",                      # 25
    "Price currency",                 # 26
    "KR_Prior_Group",                 # 27
    "KR_Prior_Group_Phase",           # 28
    "KR_Prior_Group_Correction",      # 29
]

# Combined-incident (AGREES=7, CORRECTION_FIELD=8, CORRECTION_VALUE=9, TXN_REF=12)
# No KR_Revised_Reportability column.
_KR_COMBINED = [
    "KR_Incident_Code",                           # 0
    "KR_Incident_Codes_Grouping",                 # 1
    "KR_IssueCode",                               # 2
    "KR_Diagnosis",                               # 3
    "KR_Diagnostics_key",                         # 4
    "KR_Proposed_Correction_Field",               # 5
    "KR_Proposed_Correction_Value",               # 6
    "Client_Agrees_With_KR_Proposed_Correction",  # 7  AGREES
    "Client_Final_Correction_Field",              # 8  CORRECTION_FIELD
    "Client_Final_Correction_Value",              # 9  CORRECTION_VALUE
    "KR_RECORD_KEY",                              # 10
    "Executing entity identification code",       # 11
    "Transaction reference number",               # 12 MATCH KEY
    "KR_REPORTABLE",                              # 13
    "KR_REPORTABLE_REASON",                       # 14
    "Country of the branch for the buyer",        # 15
    "Buyer identification code",                  # 16
    "Buyer_Name",                                 # 17
    "Country of the branch for the seller",       # 18
    "Seller identification code",                 # 19
    "Seller_Name",                                # 20
    "Seller decision maker code",                 # 21
    "Trading capacity",                           # 22
    "Venue",                                      # 23
    "Type of buyer identification code",          # 24
    "Type of seller identification code",         # 25
    "Type of buyer decision maker code",          # 26
    "Type of seller decision maker code",         # 27
    "KR_Prior_Group",                             # 28
    "KR_Prior_Group_Phase",                       # 29
    "KR_Prior_Group_Correction",                  # 30
]


# ===========================================================================
# 7_39  Buyer ID Validation
# Source: data/test/sample/buyer_id_validation/extract/7_39_FY26_Q1.csv
# TXN001-TXN020 (TXN008 gap, TXN020 short row)
# ===========================================================================

def _build_7_39_kr_rows() -> list[list]:
    def r(n, diag, prop_field, prop_val, agrees, txn, buyer_id, buyer_name,
          id_type="N", country="GB"):
        return [
            "7_39", "7_39", n, diag, n,
            prop_field, prop_val,
            "FALSE",
            agrees,   # 8
            None, None,  # 9, 10 written back
            n * 1000, EXEC_ENTITY,
            txn,      # 13
            True, 1,
            country, buyer_id, buyer_name,
            None, INTC, "Sample Firm", None,
            "AOTC", "XOFF",
            id_type, "I", None, None, None, None, None,
        ]

    return [
        r(1,  "Valid NIDN — no change required.",
          "Buyer identification code", "No Change",
          "Agree", "TXN001", "AB123456C", "John Smith"),
        r(2,  "NIDN suffix Z is invalid (must be A-D).",
          "Buyer identification code", "AB123456B",
          "Agree", "TXN002", "AB123456Z", "Jane Doe"),
        r(3,  "NIDN prefix BG is not a valid HMRC-issued prefix.",
          "Buyer identification code", "AB123456C",
          "Disagree", "TXN003", "BG123456C", "Robert Jones"),
        r(4,  "CONCAT ID is a deprecated format — NIDN required.",
          "Buyer identification code", "AB123456C",
          "Values Provided", "TXN004", "GB19850615JOHN#SMIT#", "John Smith",
          id_type="CONCAT"),
        r(5,  "CONCAT DOB 1990-12-25 does not match reported DOB 1985-06-15.",
          "Buyer identification code", "AB123456C",
          "False", "TXN005", "GB19901225JOHN#SMIT#", "John Smith",
          id_type="CONCAT"),
        r(6,  "Type of Buyer ID Code should be CONCAT not NIDN.",
          "Type of buyer identification code", "CONCAT",
          "Agree", "TXN006", "GB19850615JOHN#SMIT#", "John Smith",
          id_type="N"),
        r(7,  "Buyer is a legal entity with a valid LEI — no change.",
          "Buyer identification code", "No Change",
          "Agree", "TXN007", "549300TESTLEI0000001", "Test Corp Ltd",
          id_type="L", country=""),
        r(9,  "Buyer ID missing for non-UK individual (DE).",
          "Buyer identification code", "DE19880410ALICWILLI",
          "Agree", "TXN009", "", "Alice Williams",
          id_type="CONCAT", country="DE"),
        r(10, "Buyer ID missing — CONCAT required for FR national.",
          "Buyer identification code", "FR19920722BOB#TAYL#",
          "Agree", "TXN010", "", "Bob Taylor",
          id_type="CONCAT", country="FR"),
        r(11, "Buyer ID uses internal fallback prefix — NIDN required.",
          "Buyer identification code", "AB123456C",
          "Agree", "TXN011", "GB_P011", "Charlie Brown"),
        r(12, "Italian Codice Fiscale format valid — no change.",
          "Buyer identification code", "No Change",
          "Agree", "TXN012", "RSSMRA80A01H501U", "Mario Rossi",
          country="IT"),
        r(13, "Italian CF has invalid gender encoding in day field.",
          "Buyer identification code", "RSSMRA80A01H501U",
          "Agree", "TXN013", "RSSMRA80X01H501U", "Mario Rossi",
          country="IT"),
        r(14, "Swedish NIDN format valid — no change.",
          "Buyer identification code", "No Change",
          "Agree", "TXN014", "199001012389", "Erik Johansson",
          country="SE"),
        r(15, "Swedish NIDN has non-numeric check characters.",
          "Buyer identification code", "199001012389",
          "Agree", "TXN015", "1990010123XX", "Erik Johansson",
          country="SE"),
        r(16, "French CONCAT format valid — no change.",
          "Buyer identification code", "No Change",
          "Agree", "TXN016", "FR19850615JEAN#DUPON", "Jean Dupont",
          id_type="CONCAT", country="FR"),
        r(17, "Cypriot passport reviewed — valid.",
          "Buyer identification code", "No Change",
          "Agree", "TXN017", "E123456", "Maria Santos",
          id_type="CCPT", country="CY"),
        r(18, "JNT account — first holder buyer ID reviewed.",
          "Buyer identification code", "AB123456C",
          "Agree", "TXN018", "AB123456C", "John Smith"),
        r(19, "JNT account — second holder buyer ID reviewed.",
          "Buyer identification code", "CD345678E",
          "Agree", "TXN019", "CD345678E", "Jane Smith"),
        r(20, "Row is incomplete — flagged for review.",
          "Buyer identification code", "Client review",
          None, "TXN020", None, None),
    ]


def _build_7_39_template_rows() -> list[list]:
    return [
        _TMPL_HDR,
        ["TXN001", "No Change", "", "Y", "", "", ""],
        ["TXN002", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN003", "AB123456C", "Buyer identification code", "N",
         "AB123456C", "Buyer identification code", ""],
        ["TXN004", "AB123456C", "Buyer identification code", "P", "", "", ""],
        ["TXN005", "AB123456C", "Buyer identification code", "F",
         "AB123456C", "Buyer identification code", ""],
        ["TXN006", "CONCAT", "Type of buyer identification code", "Y", "", "", ""],
        ["TXN007", "No Change", "", "Y", "", "", ""],
        ["TXN009", "DE19880410ALICWILLI", "Buyer identification code", "Y", "", "", ""],
        ["TXN010", "FR19920722BOB#TAYL#", "Buyer identification code", "Y", "", "", ""],
        ["TXN011", "AB123456C", "Buyer identification code", "Y", "", "", ""],
        ["TXN012", "No Change", "", "Y", "", "", ""],
        ["TXN013", "RSSMRA80A01H501U", "Buyer identification code", "Y", "", "", ""],
        ["TXN014", "No Change", "", "Y", "", "", ""],
        ["TXN015", "199001012389", "Buyer identification code", "Y", "", "", ""],
        ["TXN016", "No Change", "", "Y", "", "", ""],
        ["TXN017", "No Change", "", "Y", "", "", ""],
        ["TXN018", "AB123456C", "Buyer identification code", "Y", "", "", ""],
        ["TXN019", "CD345678E", "Buyer identification code", "Y", "", "", ""],
        # TXN020 absent -> processor logs unmatched warning
    ]


# ===========================================================================
# 35_3  Net Amount
# Source: data/test/sample/net_amount/35_3_FY26_Q1.csv
# TXN400-TXN411
# ===========================================================================

def _build_35_3_kr_rows() -> list[list]:
    def r(n, diag, agrees, txn, net, delta):
        return [
            "35_3", "35_3", n, diag, n,
            "Net amount", "Client review",
            "FALSE",
            agrees,  # 8
            None, None,
            n * 1000, EXEC_ENTITY,
            txn,     # 13
            True, 1,
            "Equity", "Equity - UK", "GB0007980591",
            net, delta, "Unit", None, str(int(net) if net else 0),
            "MntryValAmt", str(round(net / 100, 4) if net else 0), "GBP",
            None, None, None,
        ]

    return [
        r(1, "Exact match Net=Consideration+Interest — no action.",
          "Agree", "TXN400", 1000.00, 0.00),
        r(2, "Delta 0.005 within 0.01 tolerance — no action.",
          "Agree", "TXN401", 1000.00, 0.005),
        r(3, "Delta 0.01 at tolerance boundary — review.",
          "Agree", "TXN402", 1000.00, 0.01),
        r(4, "Delta 0.02 exceeds 0.01 tolerance — correction required.",
          "Values Provided", "TXN403", 1000.00, 0.02),
        r(5, "Negative interest — net amount matches — no action.",
          "Agree", "TXN404", 980.00, 0.00),
        r(6, "Zero interest — net amount matches — no action.",
          "Agree", "TXN405", 500.00, 0.00),
        r(7, "All zero values — no action required.",
          "Agree", "TXN406", 0.00, 0.00),
        r(8, "Large trade — delta within 0.01% of net — no action.",
          "Agree", "TXN407", 1000000.00, 0.05),
        r(9, "Large trade — delta 0.003 within tolerance — no action.",
          "Agree", "TXN408", 1000000.00, 0.003),
        r(10, "Net amount could not be parsed — client to confirm.",
          None, "TXN409", 0.00, 0.00),
        r(11, "High interest — net amount matches — no action.",
          "Agree", "TXN410", 2000.00, 0.00),
        r(12, "Small amounts — net amount matches — no action.",
          "Agree", "TXN411", 0.03, 0.00),
    ]


def _build_35_3_template_rows() -> list[list]:
    return [
        _TMPL_HDR,
        ["TXN400", "No Change", "", "Y", "", "", ""],
        ["TXN401", "No Change", "", "Y", "", "", ""],
        ["TXN402", "No Change", "", "Y", "", "", ""],
        ["TXN403", "999.98", "Net amount", "P", "", "", ""],
        ["TXN404", "No Change", "", "Y", "", "", ""],
        ["TXN405", "No Change", "", "Y", "", "", ""],
        ["TXN406", "No Change", "", "Y", "", "", ""],
        ["TXN407", "No Change", "", "Y", "", "", ""],
        ["TXN408", "No Change", "", "Y", "", "", ""],
        ["TXN409", "0.00", "Net amount", "P", "", "", ""],
        ["TXN410", "No Change", "", "Y", "", "", ""],
        ["TXN411", "No Change", "", "Y", "", "", ""],
    ]


# ===========================================================================
# 7_66 + 16_20  Inconsistent Buyer + Seller ID (combined)
# Source: inconsistent_buyer_id/extract/7_66_FY26_Q1.csv  TXN200-TXN217
#         inconsistent_seller_id/extract/16_20_FY26_Q1.csv TXN300-TXN317
# Combined file -> Phase2CombinedColumns (AGREES=7, TXN_REF=12)
# ===========================================================================

def _build_7_66_16_20_kr_rows() -> list[list]:
    COMBINED = "7_66|16_20"

    def rb(n, diag, agrees, txn, buyer_id, buyer_name, id_type="N"):
        """Buyer-side row in a combined incident."""
        return [
            COMBINED, COMBINED, n, diag, n,
            "Buyer identification code", "Client review",
            agrees,  # 7 AGREES
            None, None,  # 8, 9 written back
            n * 1000, EXEC_ENTITY,
            txn,     # 12 MATCH KEY
            True, 1,
            "GB", buyer_id, buyer_name,
            None, INTC, "Sample Firm", None,
            "AOTC", "XOFF",
            id_type, "I", None, None, None, None, None,
        ]

    def rs(n, diag, agrees, txn, seller_id, seller_name, id_type="N"):
        """Seller-side row in a combined incident."""
        return [
            COMBINED, COMBINED, n, diag, n,
            "Seller identification code", "Client review",
            agrees,  # 7 AGREES
            None, None,
            n * 1000, EXEC_ENTITY,
            txn,     # 12 MATCH KEY
            True, 1,
            None, EXEC_ENTITY, "Sample Firm",
            "GB", seller_id, seller_name, None,
            "AOTC", "XOFF",
            "L", id_type, None, None, None, None, None,
        ]

    return [
        # Group A: John Smith (PABC) — 3 transactions, IDs vary
        rb(1,  "Buyer NIDN inconsistent across TXN200-TXN202.", "Agree", "TXN200", "AB123456C", "John Smith"),
        rb(2,  "AB12345ZZ is invalid (suffix Z).",              "Agree", "TXN201", "AB12345ZZ", "John Smith"),
        rb(3,  "Most recent valid ID AB123456B — correction applied.", "Agree", "TXN202", "AB123456B", "John Smith"),
        # Group B: Jane Doe (PCON) — all AB123456C, no inconsistency
        rb(4,  "NIDN consistent — no action.", "Agree", "TXN203", "AB123456C", "Jane Doe"),
        rb(5,  "NIDN consistent — no action.", "Agree", "TXN204", "AB123456C", "Jane Doe"),
        rb(6,  "NIDN consistent — no action.", "Agree", "TXN205", "AB123456C", "Jane Doe"),
        # Group C: Robert Jones (PINV) — all AB12345ZZ (all invalid)
        rb(7,  "All IDs invalid (AB12345ZZ) — correction required.", "Values Provided", "TXN206", "AB12345ZZ", "Robert Jones"),
        rb(8,  "All IDs invalid — correction required.",             "Values Provided", "TXN207", "AB12345ZZ", "Robert Jones"),
        rb(9,  "All IDs invalid — correction required.",             "Values Provided", "TXN208", "AB12345ZZ", "Robert Jones"),
        # Group D: Maria Garcia (PFIX) — BADINVALIDID then AB123456C
        rb(10, "Oldest ID invalid; newest AB123456C is valid.",       "Agree", "TXN209", "BADINVALIDID", "Maria Garcia"),
        rb(11, "Newest valid ID AB123456C applied to all.",           "Agree", "TXN210", "AB123456C",    "Maria Garcia"),
        # Group E: David Wilson (PFAL) — fallback
        rb(12, "Fallback prefix ID — NIDN required.", "Values Provided", "TXN211", "GB_PFAL", "David Wilson"),
        # Group F: Sarah Parker (PACCT) — cross-account
        rb(13, "Cross-account inconsistency for Sarah Parker.", "Agree", "TXN212", "AB123456C", "Sarah Parker"),
        rb(14, "AB12345ZZ invalid — cross-account.",            "Agree", "TXN213", "AB12345ZZ", "Sarah Parker"),
        rb(15, "AB123456B most recent valid — correction.",     "Agree", "TXN214", "AB123456B", "Sarah Parker"),
        # Group G: James Hill (PSRT) — out-of-order dates
        rb(16, "Most recent AB123456B (2026-03-10).",           "Agree", "TXN215", "AB123456B", "James Hill"),
        rb(17, "AB12345ZZ invalid.",                            "Agree", "TXN216", "AB12345ZZ", "James Hill"),
        rb(18, "AB123456A earlier — correction to AB123456B.",  "Agree", "TXN217", "AB123456A", "James Hill"),
        # Group A (seller): Helen Brown (PSABC) — AB234567D/A/D
        rs(20, "Seller NIDN varies — AB234567D valid, AB234567A may be invalid.", "Agree", "TXN300", "AB234567D", "Helen Brown"),
        rs(21, "AB234567A inconsistent.",                                          "Agree", "TXN301", "AB234567A", "Helen Brown"),
        rs(22, "Most recent AB234567D — correction applied.",                      "Agree", "TXN302", "AB234567D", "Helen Brown"),
        # Group B (seller): Tom Davis (PSCON) — all CD345678E
        rs(23, "Seller NIDN consistent — no action.", "Agree", "TXN303", "CD345678E", "Tom Davis"),
        rs(24, "Seller NIDN consistent — no action.", "Agree", "TXN304", "CD345678E", "Tom Davis"),
        rs(25, "Seller NIDN consistent — no action.", "Agree", "TXN305", "CD345678E", "Tom Davis"),
        # Group C (seller): Carol Evans (PSINV) — all CD345678Z (all invalid)
        rs(26, "All seller IDs CD345678Z invalid — correction required.", "Values Provided", "TXN306", "CD345678Z", "Carol Evans"),
        rs(27, "All seller IDs invalid.",                                  "Values Provided", "TXN307", "CD345678Z", "Carol Evans"),
        rs(28, "All seller IDs invalid.",                                  "Values Provided", "TXN308", "CD345678Z", "Carol Evans"),
        # Group D (seller): Lena Muller (PSFIX)
        rs(29, "Oldest BADSELLERBAD invalid; newest AB234567D valid.", "Agree", "TXN309", "BADSELLERBAD", "Lena Muller"),
        rs(30, "Newest AB234567D applied to all.",                     "Agree", "TXN310", "AB234567D",    "Lena Muller"),
        # Group E (seller): Paul Laurent (PSFAL) — fallback
        rs(31, "Fallback prefix ID — NIDN required.", "Values Provided", "TXN311", "GB_PSFAL", "Paul Laurent"),
        # Group F (seller): Sophie White (PSACCT) — cross-account
        rs(32, "Cross-account inconsistency for Sophie White.", "Agree", "TXN312", "AB234567D", "Sophie White"),
        rs(33, "AB234567A inconsistent.",                        "Agree", "TXN313", "AB234567A", "Sophie White"),
        rs(34, "AB234567B inconsistent.",                        "Agree", "TXN314", "AB234567B", "Sophie White"),
        # Group G (seller): Nikos Papadopoulos (PSSRT) — out-of-order
        rs(35, "Most recent AB234567D (2026-03-10).", "Agree", "TXN315", "AB234567B", "Nikos Papadopoulos"),
        rs(36, "AB234567A invalid.",                  "Agree", "TXN316", "AB234567A", "Nikos Papadopoulos"),
        rs(37, "AB234567D most recent valid.",         "Agree", "TXN317", "AB234567D", "Nikos Papadopoulos"),
    ]


def _build_7_66_template_rows() -> list[list]:
    """
    Corrections for 7_66 (inconsistent buyer ID).
    Correction = the single NIDN to use for all transactions in the group.
    """
    return [
        _TMPL_HDR,
        # Group A: John Smith — correct all to AB123456B (most recent valid)
        ["TXN200", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN201", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN202", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        # Group B: Jane Doe — consistent, no change
        ["TXN203", "No Change", "", "Y", "", "", ""],
        ["TXN204", "No Change", "", "Y", "", "", ""],
        ["TXN205", "No Change", "", "Y", "", "", ""],
        # Group C: Robert Jones — all invalid, suggested correction AB123456C
        ["TXN206", "AB123456C", "Buyer identification code", "N",
         "AB123456C", "Buyer identification code", ""],
        ["TXN207", "AB123456C", "Buyer identification code", "N",
         "AB123456C", "Buyer identification code", ""],
        ["TXN208", "AB123456C", "Buyer identification code", "N",
         "AB123456C", "Buyer identification code", ""],
        # Group D: Maria Garcia — correct to AB123456C
        ["TXN209", "AB123456C", "Buyer identification code", "Y", "", "", ""],
        ["TXN210", "AB123456C", "Buyer identification code", "Y", "", "", ""],
        # Group E: David Wilson — client provides value (fallback)
        ["TXN211", "AB123456C", "Buyer identification code", "P", "", "", ""],
        # Group F: Sarah Parker — correct all to AB123456B
        ["TXN212", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN213", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN214", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        # Group G: James Hill — correct all to AB123456B
        ["TXN215", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN216", "AB123456B", "Buyer identification code", "Y", "", "", ""],
        ["TXN217", "AB123456B", "Buyer identification code", "Y", "", "", ""],
    ]


def _build_16_20_template_rows() -> list[list]:
    """
    Corrections for 16_20 (inconsistent seller ID).
    Correction = the single NIDN to use for all transactions in the group.
    """
    return [
        _TMPL_HDR,
        # Group A: Helen Brown — correct all to AB234567D
        ["TXN300", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN301", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN302", "AB234567D", "Seller identification code", "Y", "", "", ""],
        # Group B: Tom Davis — consistent, no change
        ["TXN303", "No Change", "", "Y", "", "", ""],
        ["TXN304", "No Change", "", "Y", "", "", ""],
        ["TXN305", "No Change", "", "Y", "", "", ""],
        # Group C: Carol Evans — all invalid, suggested CD345678E
        ["TXN306", "CD345678E", "Seller identification code", "N",
         "CD345678E", "Seller identification code", ""],
        ["TXN307", "CD345678E", "Seller identification code", "N",
         "CD345678E", "Seller identification code", ""],
        ["TXN308", "CD345678E", "Seller identification code", "N",
         "CD345678E", "Seller identification code", ""],
        # Group D: Lena Muller — correct to AB234567D
        ["TXN309", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN310", "AB234567D", "Seller identification code", "Y", "", "", ""],
        # Group E: Paul Laurent — client provides value
        ["TXN311", "AB234567D", "Seller identification code", "P", "", "", ""],
        # Group F: Sophie White — correct all to AB234567D
        ["TXN312", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN313", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN314", "AB234567D", "Seller identification code", "Y", "", "", ""],
        # Group G: Nikos Papadopoulos — correct all to AB234567D
        ["TXN315", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN316", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN317", "AB234567D", "Seller identification code", "Y", "", "", ""],
    ]


def _build_16_23_template_rows() -> list[list]:
    """
    Corrections for 16_23 (seller ID validation).
    Derived from seller_id_validation/extract/16_23_FY26_Q1.csv (TXN101-TXN119).
    """
    return [
        _TMPL_HDR,
        ["TXN101", "No Change", "", "Y", "", "", ""],
        ["TXN102", "No Change", "", "Y", "", "", ""],
        # TXN103: FD123456C invalid prefix FD
        ["TXN103", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN104", "No Change", "", "Y", "", "", ""],
        # TXN105: CONCAT DOB mismatch
        ["TXN105", "AB234567D", "Seller identification code", "F",
         "AB234567D", "Seller identification code", ""],
        # TXN106: CONCAT submitted as NIDN type
        ["TXN106", "CONCAT", "Type of seller identification code", "Y", "", "", ""],
        ["TXN107", "No Change", "", "Y", "", "", ""],
        # TXN108: invalid LEI TESTBADLE2WRONGFMTYY
        ["TXN108", "549300TESTLEI0000002", "Seller identification code", "Y", "", "", ""],
        # TXN109: empty ID, Lena Muller DE
        ["TXN109", "DE19930718LENAMULL", "Seller identification code", "Y", "", "", ""],
        # TXN110: empty ID, Paul Laurent FR
        ["TXN110", "FR19870214PAULLAUR#", "Seller identification code", "Y", "", "", ""],
        # TXN111: GB_P111 fallback
        ["TXN111", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN112", "No Change", "", "Y", "", "", ""],
        # TXN113: Italian CF invalid gender
        ["TXN113", "RSSMRA80A01H501U", "Seller identification code", "Y", "", "", ""],
        ["TXN114", "No Change", "", "Y", "", "", ""],
        # TXN115: Swedish NIDN invalid
        ["TXN115", "199001012389", "Seller identification code", "Y", "", "", ""],
        ["TXN116", "No Change", "", "Y", "", "", ""],
        ["TXN117", "No Change", "", "Y", "", "", ""],
        ["TXN118", "AB234567D", "Seller identification code", "Y", "", "", ""],
        ["TXN119", "EF567890A", "Seller identification code", "Y", "", "", ""],
        # TXN120 absent -> processor logs unmatched warning
    ]


# ===========================================================================
# Phase 3 Feedback CSVs
# Persons from 7_66 (buyer) and 16_20 (seller) inconsistent ID extracts.
# Col layout:
#   [0] Reported Name & DOB  or  Reported ID
#   [1] Reported IDs         or  Reported Names & DOBs
#   [2] Suggested Correction
#   [3] Incident Types
#   [4] Incident Codes
#   [5] Totals
#   [6] Client Confirmed Correction  <- written by phase_3_processor
#   [7] Client Confirmed Correction Fields <- written by phase_3_processor
#   [8] Client Comments
# ===========================================================================

_P3_IDS_HDR = [
    "Reported Name & DOB",
    "Reported IDs",
    "Suggested Correction (Best Efforts)",
    "Incident Types",
    "Incident Codes",
    "Totals",
    "Client Confirmed Correction",
    "Client Confirmed Correction Fields",
    "Client Comments",
]

_P3_NAMES_HDR = [
    "Reported ID",
    "Reported Names & DOBs",
    "Suggested Correction (Best Efforts)",
    "Incident Types",
    "Incident Codes",
    "Totals",
    "Client Confirmed Correction",
    "Client Confirmed Correction Fields",
    "Client Comments",
]


def _build_p3_ids_feedback_rows() -> list[list]:
    """IDs Summary (before phase_3_processor runs — corrections left blank)."""
    return [
        _P3_IDS_HDR,
        # John Smith — AB123456C, AB12345ZZ, AB123456B
        ["JOHN~SMITH~1985-06-15",
         "NIDN:AB123456C\nNIDN:AB12345ZZ\nNIDN:AB123456B",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         None, None, None],
        # Jane Doe — all AB123456C (consistent)
        ["JANE~DOE~1990-03-20",
         "NIDN:AB123456C",
         None, "Inconsistent ID", "7_66", "3",
         None, None, "All transactions report same ID"],
        # Robert Jones — all AB12345ZZ (all invalid)
        ["ROBERT~JONES~1975-12-01",
         "NIDN:AB12345ZZ",
         "AB123456C", "Inconsistent ID|Invalid ID format", "7_66|7_39", "3",
         None, None, None],
        # Maria Garcia — BADINVALIDID then AB123456C
        ["MARIA~GARCIA~1988-07-22",
         "NIDN:BADINVALIDID\nNIDN:AB123456C",
         "AB123456C", "Inconsistent ID", "7_66", "2",
         None, None, None],
        # David Wilson — GB_PFAL fallback
        ["DAVID~WILSON~1982-11-30",
         "NIDN:GB_PFAL",
         None, "Inconsistent ID|Invalid ID format", "7_66|7_39", "1",
         None, None, "Fallback ID — client to confirm NIDN"],
        # Sarah Parker — cross-account AB123456C, AB12345ZZ, AB123456B
        ["SARAH~PARKER~1980-04-15",
         "NIDN:AB123456C\nNIDN:AB12345ZZ\nNIDN:AB123456B",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         None, None, None],
        # James Hill — out-of-order AB123456B, AB12345ZZ, AB123456A
        ["JAMES~HILL~1977-08-03",
         "NIDN:AB123456B\nNIDN:AB12345ZZ\nNIDN:AB123456A",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         None, None, None],
        # Person with no matching correction entry in the template
        ["UNKNOWN~PERSON~1960-01-01",
         "NIDN:AB999999A",
         None, "Inconsistent ID", "7_66", "1",
         None, None, None],
    ]


def _build_p3_names_feedback_rows() -> list[list]:
    """Names Summary (before phase_3_processor runs — corrections left blank)."""
    return [
        _P3_NAMES_HDR,
        # Helen Brown — AB234567D (valid), AB234567A (inconsistent variant)
        ["AB234567D~NIDN",
         "HELEN:BROWN:1982-09-10",
         None, "Inconsistent ID", "16_20", "2",
         None, None, "Name consistent — ID inconsistency resolved separately"],
        # Tom Davis — CD345678E consistent
        ["CD345678E~NIDN",
         "TOM:DAVIS:1978-04-25",
         None, "Inconsistent ID", "16_20", "3",
         None, None, None],
        # Carol Evans — CD345678Z all invalid
        ["CD345678Z~NIDN",
         "CAROL:EVANS:1991-11-08",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "3",
         None, None, None],
        # Helen Brown variant — AB234567A needs correcting to AB234567D
        ["AB234567A~NIDN",
         "HELEN:BROWN:1982-09-10",
         None, "Inconsistent ID", "16_20", "1",
         None, None, None],
        # Lena Muller — BADSELLERBAD
        ["BADSELLERBAD~NIDN",
         "LENA:MULLER:1993-07-18",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "1",
         None, None, None],
        # Paul Laurent — GB_PSFAL fallback
        ["GB_PSFAL~NIDN",
         "PAUL:LAURENT:1987-02-14",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "1",
         None, None, "Fallback ID — client to confirm"],
        # Sophie White — AB234567B variant
        ["AB234567B~NIDN",
         "SOPHIE:WHITE:1979-08-22",
         None, "Inconsistent ID", "16_20", "1",
         None, None, None],
        # Nikos — AB234567D most recent valid
        ["AB234567D~NIDN",
         "NIKOS:PAPADOPOULOS:1986-03-20\nHELEN:BROWN:1982-09-10",
         None, "Inconsistent Name", "13_1", "2",
         None, None, "Two persons share this ID — investigate"],
        # Person not in incident template
        ["AB999888C~NIDN",
         "UNKNOWN:PERSON:1960-01-01",
         None, "Inconsistent ID", "16_20", "1",
         None, None, None],
        # INTC counterparty
        ["INTC~",
         "INTC:INTC:",
         None, "N/A", "7_11", "5",
         None, None, "Internal to firm"],
    ]


# ===========================================================================
# Phase 3 Final Lookup — corrections pre-filled (as if phase_3_processor ran)
# ===========================================================================

def _build_p3_final_ids_rows() -> list[list]:
    """IDs Summary with corrections filled in."""
    return [
        _P3_IDS_HDR,
        ["JOHN~SMITH~1985-06-15",
         "NIDN:AB123456C\nNIDN:AB12345ZZ\nNIDN:AB123456B",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         "AB123456B", "ID", None],
        ["JANE~DOE~1990-03-20",
         "NIDN:AB123456C",
         None, "Inconsistent ID", "7_66", "3",
         "No Change", None, None],
        ["ROBERT~JONES~1975-12-01",
         "NIDN:AB12345ZZ",
         "AB123456C", "Inconsistent ID|Invalid ID format", "7_66|7_39", "3",
         "AB123456C", "ID", None],
        ["MARIA~GARCIA~1988-07-22",
         "NIDN:BADINVALIDID\nNIDN:AB123456C",
         "AB123456C", "Inconsistent ID", "7_66", "2",
         "AB123456C", "ID", None],
        # David Wilson: correction pending — remains blank (tests 'client not found' path)
        ["DAVID~WILSON~1982-11-30",
         "NIDN:GB_PFAL",
         None, "Inconsistent ID|Invalid ID format", "7_66|7_39", "1",
         None, None, None],
        ["SARAH~PARKER~1980-04-15",
         "NIDN:AB123456C\nNIDN:AB12345ZZ\nNIDN:AB123456B",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         "AB123456B", "ID", None],
        ["JAMES~HILL~1977-08-03",
         "NIDN:AB123456B\nNIDN:AB12345ZZ\nNIDN:AB123456A",
         "AB123456B", "Inconsistent ID", "7_66", "3",
         "AB123456B", "ID", None],
        ["UNKNOWN~PERSON~1960-01-01",
         "NIDN:AB999999A",
         None, "Inconsistent ID", "7_66", "1",
         None, None, None],
    ]


def _build_p3_final_names_rows() -> list[list]:
    """Names Summary with corrections filled in."""
    return [
        _P3_NAMES_HDR,
        ["AB234567D~NIDN",
         "HELEN:BROWN:1982-09-10",
         None, "Inconsistent ID", "16_20", "2",
         "No Change", None, None],
        ["CD345678E~NIDN",
         "TOM:DAVIS:1978-04-25",
         None, "Inconsistent ID", "16_20", "3",
         "No Change", None, None],
        # Carol Evans: deliberately wrong correction to exercise the FAIL path
        ["CD345678Z~NIDN",
         "CAROL:EVANS:1991-11-08",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "3",
         "CD345678WRONGVAL", "ID", None],
        ["AB234567A~NIDN",
         "HELEN:BROWN:1982-09-10",
         None, "Inconsistent ID", "16_20", "1",
         "AB234567D", "ID", None],
        ["BADSELLERBAD~NIDN",
         "LENA:MULLER:1993-07-18",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "1",
         "AB234567D", "ID", None],
        # Paul Laurent: correction pending — tests 'client not found'
        ["GB_PSFAL~NIDN",
         "PAUL:LAURENT:1987-02-14",
         None, "Inconsistent ID|Invalid ID format", "16_20|16_23", "1",
         None, None, None],
        ["AB234567B~NIDN",
         "SOPHIE:WHITE:1979-08-22",
         None, "Inconsistent ID", "16_20", "1",
         "AB234567D", "ID", None],
        ["AB234567D~NIDN",
         "NIKOS:PAPADOPOULOS:1986-03-20\nHELEN:BROWN:1982-09-10",
         None, "Inconsistent Name", "13_1", "2",
         "No Change", None, None],
        ["AB999888C~NIDN",
         "UNKNOWN:PERSON:1960-01-01",
         None, "Inconsistent ID", "16_20", "1",
         None, None, None],
        ["INTC~",
         "INTC:INTC:",
         None, "N/A", "7_11", "5",
         "No Change", None, None],
    ]


# ===========================================================================
# UnaVista CSV  (87 columns)
# Buyer IDs  = corrected values from 7_66 IDs Summary (TXN200-TXN217)
# Seller IDs = corrected values from 16_20 Names Summary (TXN300-TXN317)
# col[8]  = Buyer ID  (match key for phase_3_final_lookup IDs processor)
# col[21] = Seller ID (match key for phase_3_final_lookup Names processor)
# ===========================================================================

_UVISTA_HDR = [
    "Report Status",                        # 0
    "Transaction Reference Number",         # 1
    "Venue Transaction ID",                 # 2
    "Submitting Entity ID",                 # 3
    "Executing Entity ID",                  # 4
    "Investment Firm Indicator",            # 5
    "Buyer ID Type",                        # 6
    "Buyer ID Sub Type",                    # 7
    "Buyer ID",                             # 8  <- match key (Buyer)
    "Buyer Country of Branch",              # 9
    "Buyer First Name",                     # 10
    "Buyer Surname",                        # 11
    "Buyer DOB",                            # 12
    "Buyer Decision Maker ID Type",         # 13
    "Buyer Decision Maker ID Sub Type",     # 14
    "Buyer Decision Maker ID",              # 15
    "Buyer Decision Maker First Name",      # 16
    "Buyer Decision Maker Surname",         # 17
    "Buyer Decision Maker DOB",             # 18
    "Seller ID Type",                       # 19
    "Seller ID Sub Type",                   # 20
    "Seller ID",                            # 21  <- match key (Seller)
    "Seller Country of Branch",             # 22
    "Seller First Name",                    # 23
    "Seller Surname",                       # 24
    "Seller DOB",                           # 25
    "Seller Decision Maker ID Type",        # 26
    "Seller Decision Maker ID Sub Type",    # 27
    "Seller Decision Maker ID",             # 28
    "Seller Decision Maker First Name",     # 29
    "Seller Decision Maker Surname",        # 30
    "Seller Decision Maker DOB",            # 31
    "Order Transmission Indicator",         # 32
    "Buyer Transmitter ID",                 # 33
    "Seller Transmitter ID",                # 34
    "Trading Date Time",                    # 35
    "Trading Capacity",                     # 36
    "Quantity",                             # 37
    "Quantity Type",                        # 38
    "Quantity Currency",                    # 39
    "Derivative Notional Change",           # 40
    "Price",                                # 41
    "Price Type",                           # 42
    "Price Currency",                       # 43
    "Net Amount",                           # 44
    "Venue",                                # 45
    "Country of Branch",                    # 46
    "Up-Front Payment",                     # 47
    "Up-Front Payment Currency",            # 48
    "Complex Trade Component ID",           # 49
    "Instrument ID Type",                   # 50
    "Instrument ID",                        # 51
    "Instrument Name",                      # 52
    "Instrument Classification",            # 53
    "Notional Currency 1",                  # 54
    "Notional Currency 2",                  # 55
    "Notional Currency 2 Type",             # 56
    "Price Multiplier",                     # 57
    "UV Instrument Classification",         # 58
    "Underlying Instrument ID",             # 59
    "UV Index Classification",              # 60
    "Underlying Index ID",                  # 61
    "Underlying Index Name",                # 62
    "Underlying Index Term",                # 63
    "Option Type",                          # 64
    "Strike Price",                         # 65
    "Strike Price Type",                    # 66
    "Strike Price Currency",                # 67
    "Option Style",                         # 68
    "Maturity Date",                        # 69
    "Expiry Date",                          # 70
    "Delivery Type",                        # 71
    "Investment Decision ID Type",          # 72
    "Investment Decision ID Sub Type",      # 73
    "Investment Decision ID",               # 74
    "Investment Decision Country of Branch", # 75
    "Firm Execution ID Type",               # 76
    "Firm Execution ID Sub Type",           # 77
    "Firm Execution ID",                    # 78
    "Firm Execution Country of Branch",     # 79
    "Waiver Indicator",                     # 80
    "Short Selling Indicator",              # 81
    "OTC Post Trade Indicator",             # 82
    "Commodity Derivative Indicator",       # 83
    "SFT Indicator",                        # 84
    "Internal Client Identification",       # 85
    "Data Category",                        # 86
]
assert len(_UVISTA_HDR) == 87


def _uv_row(txn, buyer_id_type, buyer_id_sub, buyer_id,
            buyer_country, buyer_fn, buyer_sn, buyer_dob,
            seller_id_type, seller_id_sub, seller_id,
            seller_country="", seller_fn="", seller_sn="", seller_dob="",
            qty=100, price=10.0, net=1000.0) -> list:
    r = [""] * 87
    r[0] = "NEWT"
    r[1] = txn
    r[3] = EXEC_ENTITY
    r[4] = EXEC_ENTITY
    r[5] = "True"
    r[6] = buyer_id_type
    r[7] = buyer_id_sub
    r[8] = buyer_id
    r[9] = buyer_country
    r[10] = buyer_fn
    r[11] = buyer_sn
    r[12] = buyer_dob
    r[19] = seller_id_type
    r[20] = seller_id_sub
    r[21] = seller_id
    r[22] = seller_country
    r[23] = seller_fn
    r[24] = seller_sn
    r[25] = seller_dob
    r[32] = "False"
    r[35] = "2025-07-01T08:00:00.000000Z"
    r[36] = "AOTC"
    r[37] = str(qty)
    r[38] = "Unit"
    r[41] = str(price)
    r[42] = "MntryValAmt"
    r[43] = "GBP"
    r[44] = str(net)
    r[45] = "XOFF"
    r[50] = "FinInstrm.Id"
    r[51] = "GB0007980591"
    r[86] = "N"
    return r


def _build_unavista_rows() -> list[list]:
    """
    Buyer-side rows use corrected buyer IDs from 7_66 IDs Summary.
    Seller-side rows use corrected seller IDs from 16_20 Names Summary.
    """
    rows = [_UVISTA_HDR]

    # ── Buyer-side (TXN200-TXN217, corrected buyer IDs) ──────────────────────
    # Group A: John Smith — corrected to AB123456B
    rows.append(_uv_row("TXN200", "N", "NIDN", "AB123456B", "GB", "JOHN", "SMITH", "1985-06-15", "I", "", INTC, qty=150, price=6.67, net=1000.5))
    rows.append(_uv_row("TXN201", "N", "NIDN", "AB123456B", "GB", "JOHN", "SMITH", "1985-06-15", "I", "", INTC, qty=200, price=5.0,  net=1000.0))
    rows.append(_uv_row("TXN202", "N", "NIDN", "AB123456B", "GB", "JOHN", "SMITH", "1985-06-15", "I", "", INTC, qty=100, price=10.0, net=1000.0))
    # Group B: Jane Doe — AB123456C (no change)
    rows.append(_uv_row("TXN203", "N", "NIDN", "AB123456C", "GB", "JANE", "DOE", "1990-03-20", "I", "", INTC, qty=75, price=13.33, net=999.75))
    rows.append(_uv_row("TXN204", "N", "NIDN", "AB123456C", "GB", "JANE", "DOE", "1990-03-20", "I", "", INTC, qty=80, price=12.5, net=1000.0))
    rows.append(_uv_row("TXN205", "N", "NIDN", "AB123456C", "GB", "JANE", "DOE", "1990-03-20", "I", "", INTC, qty=120, price=8.33, net=999.6))
    # Group C: Robert Jones — corrected to AB123456C
    rows.append(_uv_row("TXN206", "N", "NIDN", "AB123456C", "GB", "ROBERT", "JONES", "1975-12-01", "I", "", INTC, qty=50, price=20.0, net=1000.0))
    rows.append(_uv_row("TXN207", "N", "NIDN", "AB123456C", "GB", "ROBERT", "JONES", "1975-12-01", "I", "", INTC, qty=90, price=11.11, net=999.9))
    rows.append(_uv_row("TXN208", "N", "NIDN", "AB123456C", "GB", "ROBERT", "JONES", "1975-12-01", "I", "", INTC, qty=110, price=9.09, net=999.9))
    # Group D: Maria Garcia — corrected to AB123456C
    rows.append(_uv_row("TXN209", "N", "NIDN", "AB123456C", "GB", "MARIA", "GARCIA", "1988-07-22", "I", "", INTC, qty=60, price=16.67, net=1000.2))
    rows.append(_uv_row("TXN210", "N", "NIDN", "AB123456C", "GB", "MARIA", "GARCIA", "1988-07-22", "I", "", INTC, qty=40, price=25.0, net=1000.0))
    # Group E: David Wilson — fallback still present (correction pending)
    rows.append(_uv_row("TXN211", "N", "NIDN", "GB_PFAL",   "GB", "DAVID", "WILSON", "1982-11-30", "I", "", INTC, qty=30, price=33.33, net=999.9))
    # Group F: Sarah Parker — corrected to AB123456B
    rows.append(_uv_row("TXN212", "N", "NIDN", "AB123456B", "GB", "SARAH", "PARKER", "1980-04-15", "I", "", INTC, qty=200, price=5.0, net=1000.0))
    rows.append(_uv_row("TXN213", "N", "NIDN", "AB123456B", "GB", "SARAH", "PARKER", "1980-04-15", "I", "", INTC, qty=175, price=5.71, net=999.25))
    rows.append(_uv_row("TXN214", "N", "NIDN", "AB123456B", "GB", "SARAH", "PARKER", "1980-04-15", "I", "", INTC, qty=150, price=6.67, net=1000.5))
    # Group G: James Hill — corrected to AB123456B
    rows.append(_uv_row("TXN215", "N", "NIDN", "AB123456B", "GB", "JAMES", "HILL", "1977-08-03", "I", "", INTC, qty=250, price=4.0, net=1000.0))
    rows.append(_uv_row("TXN216", "N", "NIDN", "AB123456B", "GB", "JAMES", "HILL", "1977-08-03", "I", "", INTC, qty=300, price=3.33, net=999.0))
    rows.append(_uv_row("TXN217", "N", "NIDN", "AB123456B", "GB", "JAMES", "HILL", "1977-08-03", "I", "", INTC, qty=180, price=5.56, net=1000.8))

    # ── Seller-side (TXN300-TXN317, corrected seller IDs) ────────────────────
    # Group A: Helen Brown — corrected to AB234567D
    rows.append(_uv_row("TXN300", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "HELEN", "BROWN", "1982-09-10", qty=100, price=10.0, net=1000.0))
    rows.append(_uv_row("TXN301", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "HELEN", "BROWN", "1982-09-10", qty=80,  price=12.5, net=1000.0))
    rows.append(_uv_row("TXN302", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "HELEN", "BROWN", "1982-09-10", qty=120, price=8.33, net=999.6))
    # Group B: Tom Davis — CD345678E (no change)
    rows.append(_uv_row("TXN303", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "TOM", "DAVIS", "1978-04-25", qty=60, price=16.67, net=1000.2))
    rows.append(_uv_row("TXN304", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "TOM", "DAVIS", "1978-04-25", qty=90, price=11.11, net=999.9))
    rows.append(_uv_row("TXN305", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "TOM", "DAVIS", "1978-04-25", qty=50, price=20.0, net=1000.0))
    # Group C: Carol Evans — CD345678E in UnaVista; Names Summary has wrong correction (FAIL test)
    rows.append(_uv_row("TXN306", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "CAROL", "EVANS", "1991-11-08", qty=40, price=25.0, net=1000.0))
    rows.append(_uv_row("TXN307", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "CAROL", "EVANS", "1991-11-08", qty=70, price=14.29, net=1000.3))
    rows.append(_uv_row("TXN308", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "CD345678E", "GB", "CAROL", "EVANS", "1991-11-08", qty=55, price=18.18, net=999.9))
    # Group D: Lena Muller — corrected to AB234567D
    rows.append(_uv_row("TXN309", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "LENA", "MULLER", "1993-07-18", qty=30, price=33.33, net=999.9))
    rows.append(_uv_row("TXN310", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "LENA", "MULLER", "1993-07-18", qty=45, price=22.22, net=999.9))
    # Group E: Paul Laurent — fallback still present
    rows.append(_uv_row("TXN311", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "GB_PSFAL",  "GB", "PAUL", "LAURENT", "1987-02-14", qty=25, price=40.0, net=1000.0))
    # Group F: Sophie White — corrected to AB234567D
    rows.append(_uv_row("TXN312", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "SOPHIE", "WHITE", "1979-08-22", qty=100, price=10.0, net=1000.0))
    rows.append(_uv_row("TXN313", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "SOPHIE", "WHITE", "1979-08-22", qty=130, price=7.69, net=999.7))
    rows.append(_uv_row("TXN314", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "SOPHIE", "WHITE", "1979-08-22", qty=85,  price=11.76, net=999.6))
    # Group G: Nikos Papadopoulos — corrected to AB234567D
    rows.append(_uv_row("TXN315", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "NIKOS", "PAPADOPOULOS", "1986-03-20", qty=200, price=5.0, net=1000.0))
    rows.append(_uv_row("TXN316", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "NIKOS", "PAPADOPOULOS", "1986-03-20", qty=170, price=5.88, net=999.6))
    rows.append(_uv_row("TXN317", "L", "", EXEC_ENTITY, "", "", "", "", "N", "NIDN", "AB234567D", "GB", "NIKOS", "PAPADOPOULOS", "1986-03-20", qty=220, price=4.55, net=1001.0))

    return rows


# ===========================================================================
# Orchestration
# ===========================================================================

def create_phase_2_feedback(base_dir: Path) -> None:
    out_dir = base_dir / "phase_2_feedback"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    _make_dir(out_dir)

    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_0-7~7_39~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_SINGLE_IDENTITY,
        _build_7_39_kr_rows(),
    )
    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_30-39~35_3~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_SINGLE_NET_AMOUNT,
        _build_35_3_kr_rows(),
    )
    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_0-19~7_66+16_20~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_COMBINED,
        _build_7_66_16_20_kr_rows(),
    )


def create_phase_2_incident_templates(base_dir: Path) -> None:
    out_dir = base_dir / "phase_2_incident_templates"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    _make_dir(out_dir)

    _write_csv(out_dir / "FY26 Q1 7_39.csv",  _build_7_39_template_rows())
    _write_csv(out_dir / "FY26 Q1 35_3.csv",  _build_35_3_template_rows())
    _write_csv(out_dir / "FY26 Q1 7_66.csv",  _build_7_66_template_rows())
    _write_csv(out_dir / "FY26 Q1 16_20.csv", _build_16_20_template_rows())
    _write_csv(out_dir / "FY26 Q1 16_23.csv", _build_16_23_template_rows())


def create_phase_3_feedback(base_dir: Path) -> None:
    out_dir = base_dir / "phase_3_feedback"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    _make_dir(out_dir)

    _write_csv(
        out_dir / "Replay_2025Q3_PHASE 3_Inconsistent_IDs_Summary_FINAL.csv",
        _build_p3_ids_feedback_rows(),
    )
    _write_csv(
        out_dir / "Replay_2025Q3_PHASE 3_Inconsistent_Names_Summary_FINAL.csv",
        _build_p3_names_feedback_rows(),
    )


def create_phase_3_final_lookup(base_dir: Path) -> None:
    out_dir = base_dir / "phase_3_final_lookup"
    if out_dir.exists():
        shutil.rmtree(out_dir)
    _make_dir(out_dir)

    _write_csv(
        out_dir / "Replay_2025Q3_Inconsistent_IDs_Summary_FINAL.csv",
        _build_p3_final_ids_rows(),
    )
    _write_csv(
        out_dir / "Replay_2025Q3_Inconsistent_Names_Summary_FINAL.csv",
        _build_p3_final_names_rows(),
    )
    _write_csv(
        out_dir / "UnaVista_MiFIR_Manual_Corrections_423_20180406111252.(264).csv",
        _build_unavista_rows(),
    )


def main() -> None:
    print("Generating pipeline-aligned replay sample data...")
    print(f"Output root: {BASE_DIR}\n")

    print("=== Phase 2 feedback XLSX ===")
    create_phase_2_feedback(BASE_DIR)

    print("\n=== Phase 2 incident template CSVs ===")
    create_phase_2_incident_templates(BASE_DIR)

    print("\n=== Phase 3 feedback CSVs ===")
    create_phase_3_feedback(BASE_DIR)

    print("\n=== Phase 3 final lookup files ===")
    create_phase_3_final_lookup(BASE_DIR)

    print("\nDone. All replay sample files generated.\n")
    for sub in ("phase_2_feedback", "phase_2_incident_templates",
                "phase_3_feedback", "phase_3_final_lookup"):
        d = BASE_DIR / sub
        files = sorted(d.iterdir())
        print(f"  {sub}/  ({len(files)} files)")
        for f in files:
            print(f"    {f.name}")


if __name__ == "__main__":
    main()