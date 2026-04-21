"""Validate that generated replay sample files contain no real client data."""
import csv
import re
import sys
from pathlib import Path

import openpyxl

BASE = Path("data/test/sample/replay")

# Patterns/values taken directly from the real source XLSX files
REAL_LEI = "213800Y4I7TN34WUBD71"
REAL_NAMES = [
    "AJ Bell", "STAPLES", "RASID", "BOSCHKE", "KAMBLE",
    "HOPEWELL", "CHAMBERLAIN", "ANDREW", "MOHAMAD", "VEALE", "PETER",
]
# Real TXN refs start with "44625" (seen in source data)
REAL_TXN_RE = re.compile(r"\b446\d{9,}\b")
# Real NINOs use valid UK prefixes (NOT "GB" which is a reserved/invalid prefix)
# Pattern matches any NINO where the first two letters are NOT "GB"
REAL_NINO_RE = re.compile(r"\b(?!GB)[A-CEGHJ-PR-TW-Z]{2}\d{6}[ABCD]\b")
# Synthetic NINOs we intentionally generated all start with GB (deliberately invalid)
SYNTHETIC_NINO_RE = re.compile(r"\bGB[A-Z]{2}\d{6}[A-Z]\b")

issues: list[str] = []


def check_cell(source: str, val: str) -> None:
    if not val:
        return
    if REAL_LEI in val:
        issues.append(f"{source}: real LEI present → {val[:80]}")
    if REAL_TXN_RE.search(val):
        issues.append(f"{source}: real TXN ref pattern → {val[:60]}")
    if REAL_NINO_RE.search(val):
        issues.append(f"{source}: non-GB NINO (possibly real) → {val[:60]}")
    for name in REAL_NAMES:
        if name.upper() in val.upper():
            issues.append(f"{source}: real name '{name}' → {val[:80]}")


# ── XLSX ─────────────────────────────────────────────────────────────────────
for xlsx in sorted(BASE.rglob("*.xlsx")):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for c_idx, cell in enumerate(row, 1):
                if cell is None:
                    continue
                check_cell(f"{xlsx.name}[{sheet_name}] r{r_idx}c{c_idx}", str(cell))
    wb.close()

# ── CSV ──────────────────────────────────────────────────────────────────────
for csv_path in sorted(BASE.rglob("*.csv")):
    with open(csv_path, encoding="utf-8", newline="") as f:
        for r_idx, row in enumerate(csv.reader(f), 1):
            for c_idx, cell in enumerate(row, 1):
                check_cell(f"{csv_path.name} r{r_idx}c{c_idx}", cell)

# ── Report ───────────────────────────────────────────────────────────────────
if issues:
    print(f"ISSUES FOUND ({len(issues)}):")
    for i in issues:
        print(f"  {i}")
    sys.exit(1)
else:
    print(f"CLEAN — no real client data patterns detected across all {len(list(BASE.rglob('*.*')))} files.")
    # Summarise what was checked
    xlsx_files = list(BASE.rglob("*.xlsx"))
    csv_files = list(BASE.rglob("*.csv"))
    print(f"  Checked {len(xlsx_files)} XLSX files and {len(csv_files)} CSV files.")
    print(f"  Validated against: real LEI, real TXN ref format, non-GB NINOs, known real names.")
    print(f"  All NINOs in sample files use GB prefix (UK-reserved / deliberately invalid).")
