#!/usr/bin/env python3
"""
Generate Replay Sample Data
============================

Creates comprehensive sample XLSX and CSV files for testing all four replay
scripts via the web interface.  All data is synthetic — no real client
information is used.

File layout produced
--------------------
data/test/sample/replay/
├── phase_2_feedback/
│   ├── 1903a~G15~P2_0-7~7_3~KR Final Analysis_Data_1 OF 1.xlsx
│   ├── 1903a~G15~P2_30-39~35_3~KR Final Analysis_Data_1 OF 1.xlsx
│   └── 1903a~G15~P2_8-19~12_75+21_75~KR Final Analysis_Data_1 OF 1.xlsx
├── phase_2_incident_templates/
│   ├── FY26 Q1 7_3.csv
│   └── FY26 Q1 35_3.csv
├── phase_3_feedback/
│   ├── Replay_2025Q3_PHASE 3_Inconsistent_IDs_Summary_FINAL.csv
│   └── Replay_2025Q3_PHASE 3_Inconsistent_Names_Summary_FINAL.csv
└── phase_3_final_lookup/
    ├── Replay_2025Q3_Inconsistent_IDs_Summary_FINAL.csv
    ├── Replay_2025Q3_Inconsistent_Names_Summary_FINAL.csv
    └── UnaVista_MiFIR_Manual_Corrections_423_20180406111252.(264).csv

Usage
-----
    python data/test/sample/replay/generate_replay_sample_data.py
"""

import csv
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

# ---------------------------------------------------------------------------
# Output directory (relative to repo root)
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent


# ===========================================================================
# Helper utilities
# ===========================================================================

def _make_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _write_csv(filepath: Path, rows: list[list]) -> None:
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"  Created: {filepath.name}  ({len(rows) - 1} data rows)")


def _pad_row(row: list, total_cols: int) -> list:
    """Extend row with None values to reach *total_cols*."""
    return row + [None] * (total_cols - len(row))


# ===========================================================================
# Phase 2 KR Final Analysis XLSX files
# ===========================================================================

# Common header columns 0-15 (shared by all KR Analysis file types)
_KR_HEADER_BASE = [
    "KR_Incident_Code",
    "KR_Incident_Codes_Grouping",
    "KR_IssueCode",
    "KR_Diagnosis",
    "KR_Diagnostics_key",
    "KR_Proposed_Correction_Field",
    "KR_Proposed_Correction_Value",
    "KR_Revised_Reportability",
    "Client_Agrees_With_KR_Proposed_Correction",  # col 8  — AGREES (written by processor)
    "Client_Final_Correction_Field",               # col 9  — CORRECTION_FIELD (written)
    "Client_Final_Correction_Value",               # col 10 — CORRECTION_VALUE (written)
    "KR_RECORD_KEY",
    "Executing entity identification code",
    "Transaction reference number",                # col 13 — MATCH KEY
    "KR_REPORTABLE",
    "KR_REPORTABLE_REASON",
]

# Additional columns for standard buyer/seller identity incidents (e.g. 7_3)
_KR_HEADER_IDENTITY = _KR_HEADER_BASE + [
    "Country of the branch for the buyer",  # col 16
    "Buyer identification code",            # col 17
    "Buyer_Name",                           # col 18
    "Country of the branch for the seller", # col 19
    "Seller identification code",           # col 20
    "Seller_Name",                          # col 21
    "Seller decision maker code",           # col 22
    "Trading capacity",                     # col 23
    "Venue",                                # col 24
    "Type of buyer identification code",    # col 25
    "Type of seller identification code",   # col 26
    "Type of buyer decision maker code",    # col 27
    "Type of seller decision maker code",   # col 28
    "KR_Prior_Group",                       # col 29
    "KR_Prior_Group_Phase",                 # col 30
    "KR_Prior_Group_Correction",            # col 31
]

# Additional columns for net amount incidents (e.g. 35_3)
_KR_HEADER_NET_AMOUNT = _KR_HEADER_BASE + [
    "KR_Reportable_Description",    # col 16
    "KR_Instrument_Classification", # col 17
    "Instrument identification code", # col 18
    "Net amount",                   # col 19
    "Net_amount_delta",             # col 20
    "Type of quantity",             # col 21
    "Quantity currency",            # col 22
    "Quantity_str",                 # col 23
    "Type of price",                # col 24
    "Price_str",                    # col 25
    "Price currency",               # col 26
    "KR_Prior_Group",               # col 27
    "KR_Prior_Group_Phase",         # col 28
    "KR_Prior_Group_Correction",    # col 29
]

# Combined incident header — column order shifts (KR_Revised_Reportability absent)
# Phase2CombinedColumns: AGREES=7, CORRECTION_FIELD=8, CORRECTION_VALUE=9, TXN_REF=12
_KR_HEADER_COMBINED = [
    "KR_Incident_Code",               # col 0
    "KR_Incident_Codes_Grouping",     # col 1
    "KR_IssueCode",                   # col 2
    "KR_Diagnosis",                   # col 3
    "KR_Diagnostics_key",             # col 4
    "KR_Proposed_Correction_Field",   # col 5
    "KR_Proposed_Correction_Value",   # col 6
    "Client_Agrees_With_KR_Proposed_Correction",  # col 7  — AGREES
    "Client_Final_Correction_Field",               # col 8  — CORRECTION_FIELD
    "Client_Final_Correction_Value",               # col 9  — CORRECTION_VALUE
    "KR_RECORD_KEY",                               # col 10
    "Executing entity identification code",        # col 11
    "Transaction reference number",                # col 12 — MATCH KEY
    "KR_REPORTABLE",                               # col 13
    "KR_REPORTABLE_REASON",                        # col 14
    "Country of the branch for the buyer",         # col 15
    "Buyer identification code",                   # col 16
    "Buyer_Name",                                  # col 17
    "Country of the branch for the seller",        # col 18
    "Seller identification code",                  # col 19
    "Seller_Name",                                 # col 20
    "Seller decision maker code",                  # col 21
    "Trading capacity",                            # col 22
    "Venue",                                       # col 23
    "Type of buyer identification code",           # col 24
    "Type of seller identification code",          # col 25
    "Type of buyer decision maker code",           # col 26
    "Type of seller decision maker code",          # col 27
    "KR_Prior_Group",                              # col 28
    "KR_Prior_Group_Phase",                        # col 29
    "KR_Prior_Group_Correction",                   # col 30
]

TOTAL_KR_COLS = 104  # Real files have 104 columns total


def _write_kr_xlsx(
    filepath: Path,
    header: list,
    data_rows: list[list],
    sheet_name: str = "in1",
) -> None:
    """Write a KR Final Analysis xlsx file with the standard two-sheet layout."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: in1 ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = sheet_name

    # Header row — bold, light blue fill
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    padded_header = _pad_row(header, TOTAL_KR_COLS)
    ws.append(padded_header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows
    for row in data_rows:
        ws.append(_pad_row(row, TOTAL_KR_COLS))

    # ── Sheet 2: List (dropdown source) ──────────────────────────────────────
    ws_list = wb.create_sheet("List")
    ws_list.append(["List"])
    ws_list.append(["Yes"])
    ws_list.append(["No"])

    wb.save(filepath)
    print(f"  Created: {filepath.name}  ({len(data_rows)} data rows)")


# ---------------------------------------------------------------------------
# 7_3 — Buyer identity incident  (single incident, Phase2SingleColumns)
# Edge cases:
#   - Client agrees with KR proposed correction
#   - Client disagrees → uses Suggested Correction
#   - Client provides own correction ("Values Provided")
#   - No Change required
#   - Transaction ref NOT in incident template → unmatched
#   - Seller is INTC (internal counterparty)
#   - Buyer is a company (LEI)
#   - Combined incident code grouping
#   - CANC reportability scenario
#   - Empty agrees cell (blank, not yet responded)
#   - Row with dash-separated incident code grouping
# ---------------------------------------------------------------------------

def _build_7_3_rows() -> list[list]:
    """Build data rows for the 7_3 KR Final Analysis sample file."""

    def row(
        code, grouping, issue, diagnosis, diag_key,
        prop_field, prop_value, reportability,
        agrees, corr_field, corr_value,
        rec_key, exec_entity, txn_ref,
        reportable, reportable_reason,
        buyer_country, buyer_id, buyer_name,
        seller_country, seller_id, seller_name,
        seller_dm, trading, venue, buyer_id_type, seller_id_type,
        buyer_dm_type, seller_dm_type,
        prior_grp, prior_phase, prior_corr,
    ):
        return [
            code, grouping, issue, diagnosis, diag_key,
            prop_field, prop_value, reportability,
            agrees, corr_field, corr_value,
            rec_key, exec_entity, txn_ref,
            reportable, reportable_reason,
            buyer_country, buyer_id, buyer_name,
            seller_country, seller_id, seller_name,
            seller_dm, trading, venue, buyer_id_type, seller_id_type,
            buyer_dm_type, seller_dm_type,
            prior_grp, prior_phase, prior_corr,
        ]

    LEI = "213800SAMPLE0001LEI1"
    INTC = "INTC"

    return [
        # 1. Client agrees with KR correction
        row("7_3", "7_3|7_51", 1,
            "Executing entity is the buyer — should never occur for AOTC trades.",
            1, "Buyer identification code", "GBBJ112233A", "FALSE",
            "Agree", None, None,
            1001, LEI, "SAMPLE_P2_7_3_001",
            True, 1, None, LEI, "Sample Co", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "L", "I", None, None, None, None, None),

        # 2. Client disagrees → processor should apply Suggested Correction
        row("7_3", "7_3|7_51", 2,
            "Executing entity is the buyer — AOTC trade, buyer appears incorrect.",
            3, "Buyer identification code", "GBAJ223344B", "FALSE",
            "Disagree", None, None,
            1002, LEI, "SAMPLE_P2_7_3_002",
            True, 1, None, INTC, "Sample Co", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "I", "I", None, None, None, None, None),

        # 3. Client provides own correction ("Values Provided")
        row("7_3", "7_3", 5,
            "Buyer identification code missing — client to provide.",
            2, "Buyer identification code", "Client review", "FALSE",
            "Values Provided", None, None,
            1003, LEI, "SAMPLE_P2_7_3_003",
            True, 1, "GB", None, "Anon Person", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 4. No change required
        row("7_3", "7_3", 8,
            "Buyer is correctly identified — no action needed.",
            4, "Buyer identification code", "No Change", "TRUE",
            "Agree", None, None,
            1004, LEI, "SAMPLE_P2_7_3_004",
            True, 1, "GB", "GBNC445566C", "Sample Person", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 5. Transaction ref NOT in incident template → processor logs unmatched
        row("7_3", "7_3", 9,
            "Buyer identification code missing.",
            5, "Buyer identification code", "Client review", "FALSE",
            None, None, None,
            1005, LEI, "SAMPLE_P2_7_3_NOTFOUND",
            True, 1, None, None, None, None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "I", "I", None, None, None, None, None),

        # 6. INTC seller — standard
        row("7_3", "7_3", 10,
            "Buyer is not the executing entity — review required.",
            6, "Buyer identification code", "GBEJ334455D", "FALSE",
            "Agree", None, None,
            1006, LEI, "SAMPLE_P2_7_3_006",
            True, 1, None, LEI, "Sample Co", None, "INTC", "INTC Counterparty",
            None, "AOTC", "XOFF", "L", "I", None, None, None, None, None),

        # 7. LEI buyer (corporate)
        row("7_3", "7_3", 11,
            "Buyer is a legal entity but reported with individual ID type.",
            7, "Type of buyer identification code", "LEI", "FALSE",
            "Agree", None, None,
            1007, LEI, "SAMPLE_P2_7_3_007",
            True, 1, None, "549300SAMPLECORP01XY", "Corporate Buyer",
            None, INTC, "Sample Co",
            None, "DEAL", "XOFF", "N", "I", None, None, None, None, None),

        # 8. Multi-incident grouping (7_3 and 7_51 combined diagnosis)
        row("7_3", "7_3|7_51", 12,
            "Both 7_3 and 7_51 apply — buyer is the executing entity and code missing.",
            8, "Buyer identification code", "GBKL556677E", "FALSE",
            "Values Provided", None, None,
            1008, LEI, "SAMPLE_P2_7_3_008",
            True, 1, "GB", None, "Missing Buyer", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "I", "I", None, None, None, None, None),

        # 9. Blank agrees cell — not yet filled in by client
        row("7_3", "7_3", 14,
            "Buyer identification code disputed by client — awaiting response.",
            9, "Buyer identification code", "GBMN667788F", "FALSE",
            None, None, None,
            1009, LEI, "SAMPLE_P2_7_3_009",
            True, 1, "GB", "GBMN667788F", "Pending Person", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 10. Prior group reference — correction carried forward from earlier phase
        row("7_3", "7_3", 15,
            "Buyer identification code was corrected in a previous phase.",
            10, "Buyer identification code", "GBPQ778899G", "FALSE",
            "Agree", None, None,
            1010, LEI, "SAMPLE_P2_7_3_010",
            True, 1, "GB", "GBPQ778899G", "Prior Corrected", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None,
            "7_3", "Phase 1", "GBPQ778899G_OLD"),

        # 11. Seller ID incident (7_3 applied to seller side)
        row("7_3", "7_3", 16,
            "Seller identification code incorrect — review required.",
            11, "Seller identification code", "GBRS889900H", "FALSE",
            "Agree", None, None,
            1011, LEI, "SAMPLE_P2_7_3_011",
            True, 1, None, INTC, "Sample Co", "GB", "GBRS889900H", "Sample Seller",
            None, "AOTC", "XOFF", "I", "N", None, None, None, None, None),

        # 12. CANC reportability — row marked as non-reportable
        row("7_3", "7_3", 17,
            "Transaction is being cancelled — correction will apply to CANC record.",
            12, "Buyer identification code", "GBTV990011I", "FALSE",
            "Agree", None, None,
            1012, LEI, "SAMPLE_P2_7_3_012",
            False, 2, None, "GBTV990011I", "Cancel Person", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 13. Row where correction is "Partially Agree" (P value)
        row("7_3", "7_3", 18,
            "Partial agreement — client accepts field but proposes different value.",
            13, "Buyer identification code", "GBUW001122J", "FALSE",
            "Partially Agree", None, None,
            1013, LEI, "SAMPLE_P2_7_3_013",
            True, 1, "GB", "GBUW001122J", "Partial Person", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 14. Row where incident template has Suggested Correction (N/F flow)
        row("7_3", "7_3", 19,
            "Buyer identification code format invalid.",
            14, "Buyer identification code", "GBVX112233K", "FALSE",
            "False", None, None,
            1014, LEI, "SAMPLE_P2_7_3_014",
            True, 1, "GB", "BADFORMAT", "Format Issue", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None, None, None, None),

        # 15. Decision maker correction (buyer DM)
        row("7_3", "7_3", 20,
            "Buyer decision maker code missing.",
            15, "Buyer decision maker code", "GBDM223344L", "FALSE",
            "Agree", None, None,
            1015, LEI, "SAMPLE_P2_7_3_015",
            True, 1, "GB", "GBYZ334455M", "DM Person", None, INTC, "Sample Co",
            "GBDM223344L", "AOTC", "XOFF", "N", "I", "N", None, None, None, None),
    ]


# ---------------------------------------------------------------------------
# 35_3 — Net amount incident  (single incident, Phase2SingleColumns)
# ---------------------------------------------------------------------------

def _build_35_3_rows() -> list[list]:
    """Build data rows for the 35_3 KR Final Analysis sample file."""
    LEI = "213800SAMPLE0001LEI1"

    def row(
        code, grouping, issue, diagnosis, diag_key,
        prop_field, prop_value, reportability,
        agrees, corr_field, corr_value,
        rec_key, exec_entity, txn_ref,
        reportable, reportable_reason,
        kr_desc, kr_instr, instr_id,
        net_amount, delta, qty_type, qty_ccy, qty_str,
        price_type, price_str, price_ccy,
        prior_grp, prior_phase, prior_corr,
    ):
        return [
            code, grouping, issue, diagnosis, diag_key,
            prop_field, prop_value, reportability,
            agrees, corr_field, corr_value,
            rec_key, exec_entity, txn_ref,
            reportable, reportable_reason,
            kr_desc, kr_instr, instr_id,
            net_amount, delta, qty_type, qty_ccy, qty_str,
            price_type, price_str, price_ccy,
            prior_grp, prior_phase, prior_corr,
        ]

    return [
        # 1. Client provides corrected net amount
        row("35_3", "35_3|35_8", 1,
            "Net amount is more than 0.01% away from expected value.",
            1, "Net amount", "Client review", "FALSE",
            "Values Provided", None, None,
            2001, LEI, "SAMPLE_P2_35_3_001",
            True, 1,
            "Equity", "Equity — UK", "GB0007980591",
            1234.56, 0.02, "Unit", None, "100",
            "MntryValAmt", "12.3456", "GBP",
            None, None, None),

        # 2. No change — net amount difference is within tolerance
        row("35_3", "35_3", 2,
            "Net amount difference within tolerance — no action required.",
            2, "Net amount", "No Change", "TRUE",
            "Agree", None, None,
            2002, LEI, "SAMPLE_P2_35_3_002",
            True, 1,
            "Equity", "Equity — US", "US4592001014",
            2500.00, 0.001, "Unit", None, "50",
            "MntryValAmt", "50.0", "GBP",
            None, None, None),

        # 3. Disagree — suggests alternative value
        row("35_3", "35_3", 3,
            "Net amount materially different from reported value.",
            3, "Net amount", "3500.00", "FALSE",
            "Disagree", None, None,
            2003, LEI, "SAMPLE_P2_35_3_003",
            True, 1,
            "Equity", "Equity — EU", "DE0005140008",
            3499.85, 0.15, "Unit", None, "200",
            "MntryValAmt", "17.4999", "EUR",
            None, None, None),

        # 4. Transaction ref NOT in incident template → unmatched
        row("35_3", "35_3", 4,
            "Net amount not reported.",
            4, "Net amount", "Client review", "FALSE",
            None, None, None,
            2004, LEI, "SAMPLE_P2_35_3_NOTFOUND",
            True, 1,
            "Equity", "Equity — UK", "GB0031348658",
            None, None, "Unit", None, "75",
            "MntryValAmt", "20.0", "GBP",
            None, None, None),

        # 5. Partially agree
        row("35_3", "35_3|35_10", 5,
            "Net amount slightly off — client partially agrees.",
            5, "Net amount", "5000.00", "FALSE",
            "Partially Agree", None, None,
            2005, LEI, "SAMPLE_P2_35_3_005",
            True, 1,
            "Equity", "Equity — UK", "GB0005405286",
            4998.75, 1.25, "Unit", None, "250",
            "MntryValAmt", "20.0", "GBP",
            None, None, None),

        # 6. Prior group correction
        row("35_3", "35_3", 6,
            "Net amount corrected in previous phase.",
            6, "Net amount", "1000.00", "FALSE",
            "Agree", None, None,
            2006, LEI, "SAMPLE_P2_35_3_006",
            True, 1,
            "Equity", "Equity — UK", "GB0002634946",
            1000.00, 0.00, "Unit", None, "100",
            "MntryValAmt", "10.0", "GBP",
            "35_3", "Phase 1", "999.75"),

        # 7. Negative net amount (short sale scenario)
        row("35_3", "35_3", 7,
            "Net amount is negative — client to confirm.",
            7, "Net amount", "Client review", "FALSE",
            "Values Provided", None, None,
            2007, LEI, "SAMPLE_P2_35_3_007",
            True, 1,
            "Equity", "Equity — UK", "GB0004544929",
            -875.50, 0.03, "Unit", None, "50",
            "MntryValAmt", "-17.51", "GBP",
            None, None, None),

        # 8. Zero net amount edge case
        row("35_3", "35_3", 8,
            "Net amount is zero — trade may not need reporting.",
            8, "Net amount", "0.00", "FALSE",
            "Agree", None, None,
            2008, LEI, "SAMPLE_P2_35_3_008",
            True, 1,
            "Equity", "Equity — UK", "GB0005603229",
            0.00, 0.00, "Unit", None, "0",
            "MntryValAmt", "0.0", "GBP",
            None, None, None),

        # 9. False agree (F value) → uses Suggested Correction
        row("35_3", "35_3", 9,
            "Net amount is more than 0.01% off.",
            9, "Net amount", "750.00", "FALSE",
            "False", None, None,
            2009, LEI, "SAMPLE_P2_35_3_009",
            True, 1,
            "Equity", "Equity — US", "US0231351067",
            748.25, 1.75, "Unit", None, "150",
            "MntryValAmt", "4.9883", "USD",
            None, None, None),

        # 10. Blank agrees — awaiting client response
        row("35_3", "35_3", 10,
            "Net amount delta exceeds reporting threshold.",
            10, "Net amount", "12000.00", "FALSE",
            None, None, None,
            2010, LEI, "SAMPLE_P2_35_3_010",
            True, 1,
            "Equity", "Equity — UK", "GB0004082847",
            11985.00, 15.00, "Unit", None, "500",
            "MntryValAmt", "23.97", "GBP",
            None, None, None),
    ]


# ---------------------------------------------------------------------------
# 12_75+21_75 — Combined incident  (Phase2CombinedColumns)
# Combined files: AGREES=col7, CORRECTION_FIELD=col8, CORRECTION_VALUE=col9, TXN_REF=col12
# ---------------------------------------------------------------------------

def _build_combined_rows() -> list[list]:
    """Build data rows for the 12_75+21_75 combined incident sample file."""
    LEI = "213800SAMPLE0001LEI1"

    def row(
        code, grouping, issue, diagnosis, diag_key,
        prop_field, prop_value,
        agrees, corr_field, corr_value,           # col 7, 8, 9
        rec_key, exec_entity, txn_ref,            # col 10, 11, 12
        reportable, reportable_reason,
        buyer_country, buyer_id, buyer_name,
        seller_country, seller_id, seller_name,
        seller_dm, trading, venue, buyer_id_type, seller_id_type,
        buyer_dm_type, seller_dm_type,
        prior_grp, prior_phase, prior_corr,
    ):
        return [
            code, grouping, issue, diagnosis, diag_key,
            prop_field, prop_value,
            agrees, corr_field, corr_value,
            rec_key, exec_entity, txn_ref,
            reportable, reportable_reason,
            buyer_country, buyer_id, buyer_name,
            seller_country, seller_id, seller_name,
            seller_dm, trading, venue, buyer_id_type, seller_id_type,
            buyer_dm_type, seller_dm_type,
            prior_grp, prior_phase, prior_corr,
        ]

    INTC = "INTC"

    return [
        # 1. Agree — combined buyer DM and seller DM correction
        row("12_75|21_75", "12_75+21_75", 1,
            "Both buyer DM (12_75) and seller DM (21_75) codes are invalid.",
            1, "Buyer decision maker code", "GBDM_COMBINED1",
            "Agree", None, None,
            3001, LEI, "SAMPLE_P2_COMB_001",
            True, 1,
            "GB", "GBBX223344A", "Combined Buyer", None, INTC, "Sample Co",
            "GBDM_COMBINED1", "AOTC", "XOFF", "N", "I", "N", None,
            None, None, None),

        # 2. Disagree — processor uses suggested correction
        row("12_75|21_75", "12_75+21_75", 2,
            "Both buyer DM and seller DM codes missing.",
            2, "Buyer decision maker code", "GBDM_KR_SUGGEST",
            "Disagree", None, None,
            3002, LEI, "SAMPLE_P2_COMB_002",
            True, 1,
            "GB", "GBBY334455B", "Combined Buyer 2", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None,
            None, None, None),

        # 3. Values Provided
        row("12_75|21_75", "12_75+21_75", 3,
            "Client to provide DM codes for both buyer and seller.",
            3, "Buyer decision maker code", "Client review",
            "Values Provided", None, None,
            3003, LEI, "SAMPLE_P2_COMB_003",
            True, 1,
            "GB", "GBCZ445566C", "Combined Buyer 3", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None,
            None, None, None),

        # 4. No change
        row("12_75|21_75", "12_75+21_75", 4,
            "DM codes are present and valid — no change required.",
            4, "Buyer decision maker code", "No Change",
            "Agree", None, None,
            3004, LEI, "SAMPLE_P2_COMB_004",
            True, 1,
            "GB", "GBDA556677D", "Combined Buyer 4", None, INTC, "Sample Co",
            "GBDM_VALID01", "AOTC", "XOFF", "N", "I", "N", None,
            None, None, None),

        # 5. Not in incident template → unmatched
        row("12_75|21_75", "12_75+21_75", 5,
            "Combined incident — transaction not found in incident template.",
            5, "Buyer decision maker code", "Client review",
            None, None, None,
            3005, LEI, "SAMPLE_P2_COMB_NOTFOUND",
            True, 1,
            "GB", None, None, None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "I", "I", None, None,
            None, None, None),

        # 6. False agree → Suggested Correction fallback
        row("12_75|21_75", "12_75+21_75", 6,
            "DM code format incorrect.",
            6, "Buyer decision maker code", "GBDM_WRONG",
            "False", None, None,
            3006, LEI, "SAMPLE_P2_COMB_006",
            True, 1,
            "GB", "GBEB667788E", "Combined Buyer 6", None, INTC, "Sample Co",
            "BADFORMAT", "AOTC", "XOFF", "N", "I", "N", None,
            None, None, None),

        # 7. Blank agrees — pending
        row("12_75|21_75", "12_75+21_75", 7,
            "Combined DM correction — awaiting client response.",
            7, "Buyer decision maker code", "GBDM_PENDING",
            None, None, None,
            3007, LEI, "SAMPLE_P2_COMB_007",
            True, 1,
            "GB", "GBFC778899F", "Combined Buyer 7", None, INTC, "Sample Co",
            None, "AOTC", "XOFF", "N", "I", None, None,
            None, None, None),

        # 8. Partially agree
        row("12_75|21_75", "12_75+21_75", 8,
            "Client partially agrees with combined DM correction.",
            8, "Buyer decision maker code", "GBDM_PARTIAL",
            "Partially Agree", None, None,
            3008, LEI, "SAMPLE_P2_COMB_008",
            True, 1,
            "GB", "GBGD889900G", "Combined Buyer 8", None, INTC, "Sample Co",
            "GBDM_PARTIAL", "AOTC", "XOFF", "N", "I", "N", None,
            None, None, None),
    ]


def create_phase_2_feedback(base_dir: Path) -> None:
    out_dir = base_dir / "phase_2_feedback"
    _make_dir(out_dir)

    # 1. 7_3 — standard identity incident
    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_0-7~7_3~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_HEADER_IDENTITY,
        _build_7_3_rows(),
    )

    # 2. 35_3 — net amount incident
    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_30-39~35_3~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_HEADER_NET_AMOUNT,
        _build_35_3_rows(),
    )

    # 3. 12_75+21_75 — combined incident
    _write_kr_xlsx(
        out_dir / "1903a~G15~P2_8-19~12_75+21_75~KR Final Analysis_Data_1 OF 1.xlsx",
        _KR_HEADER_COMBINED,
        _build_combined_rows(),
    )


# ===========================================================================
# Phase 2 incident template CSVs  (source of corrections for Phase 2 processor)
# ===========================================================================

# These are the Kaizen accuracy-testing output CSVs the Phase 2 processor reads.
# Column names are configurable; these use the default "incident_columns" values.

_INCIDENT_TEMPLATE_HEADER = [
    "Transaction Reference",
    "Correction",
    "Correction Field",
    "Agree With Correction",
    "Suggested Correction",
    "Suggested Correction Field",
    "Error",
]


def create_phase_2_incident_templates(base_dir: Path) -> None:
    out_dir = base_dir / "phase_2_incident_templates"
    _make_dir(out_dir)

    # ── FY26 Q1 7_3.csv ──────────────────────────────────────────────────────
    rows_7_3 = [_INCIDENT_TEMPLATE_HEADER] + [
        # Matches row 1 in KR Analysis: agree → correction applied
        ["SAMPLE_P2_7_3_001", "GBBJ112233A", "Buyer identification code", "Y",
         "", "", ""],
        # Matches row 2: disagree (N) → processor uses Suggested Correction
        ["SAMPLE_P2_7_3_002", "", "Buyer identification code", "N",
         "GBAJ_SUGGESTED", "Buyer identification code", ""],
        # Matches row 3: Values Provided (P) → correction applied
        ["SAMPLE_P2_7_3_003", "GBCLIENT003", "Buyer identification code", "P",
         "", "", ""],
        # Matches row 4: no change
        ["SAMPLE_P2_7_3_004", "No Change", "", "Y", "", "", ""],
        # Row 5 NOT present → processor logs unmatched for SAMPLE_P2_7_3_NOTFOUND
        # Matches row 6: agree, INTC scenario
        ["SAMPLE_P2_7_3_006", "GBEJ334455D", "Buyer identification code", "Y",
         "", "", ""],
        # Matches row 7: LEI buyer type correction
        ["SAMPLE_P2_7_3_007", "LEI", "Type of buyer identification code", "Y",
         "", "", ""],
        # Matches row 8: multi-incident grouping
        ["SAMPLE_P2_7_3_008", "GBKL556677E", "Buyer identification code", "Y",
         "", "", ""],
        # Row 9: blank agrees in KR file → template still has correction
        ["SAMPLE_P2_7_3_009", "GBMN667788F", "Buyer identification code", "Y",
         "", "", ""],
        # Matches row 10: prior group — agree
        ["SAMPLE_P2_7_3_010", "GBPQ778899G", "Buyer identification code", "Y",
         "", "", ""],
        # Matches row 11: seller ID correction
        ["SAMPLE_P2_7_3_011", "GBRS889900H", "Seller identification code", "Y",
         "", "", ""],
        # Matches row 12: CANC scenario
        ["SAMPLE_P2_7_3_012", "GBTV990011I", "Buyer identification code", "Y",
         "", "", ""],
        # Matches row 13: Partially Agree (P)
        ["SAMPLE_P2_7_3_013", "GBUW001122J", "Buyer identification code", "P",
         "", "", ""],
        # Matches row 14: False (F) → use Suggested Correction
        ["SAMPLE_P2_7_3_014", "BADFORMAT", "Buyer identification code", "F",
         "GBVX112233K", "Buyer identification code", ""],
        # Matches row 15: DM correction
        ["SAMPLE_P2_7_3_015", "GBDM223344L", "Buyer decision maker code", "Y",
         "", "", ""],
    ]
    _write_csv(out_dir / "FY26 Q1 7_3.csv", rows_7_3)

    # ── FY26 Q1 35_3.csv ─────────────────────────────────────────────────────
    rows_35_3 = [_INCIDENT_TEMPLATE_HEADER] + [
        ["SAMPLE_P2_35_3_001", "1234.56", "Net amount", "P", "", "", ""],
        ["SAMPLE_P2_35_3_002", "No Change", "", "Y", "", "", ""],
        ["SAMPLE_P2_35_3_003", "3500.00", "Net amount", "N",
         "3499.85", "Net amount", ""],
        # SAMPLE_P2_35_3_NOTFOUND not present → unmatched
        ["SAMPLE_P2_35_3_005", "5000.00", "Net amount", "P", "", "", ""],
        ["SAMPLE_P2_35_3_006", "1000.00", "Net amount", "Y", "", "", ""],
        ["SAMPLE_P2_35_3_007", "-875.50", "Net amount", "P", "", "", ""],
        ["SAMPLE_P2_35_3_008", "0.00", "Net amount", "Y", "", "", ""],
        ["SAMPLE_P2_35_3_009", "750.00", "Net amount", "F",
         "748.25", "Net amount", ""],
        # SAMPLE_P2_35_3_010 absent (blank agrees) → processor skips
    ]
    _write_csv(out_dir / "FY26 Q1 35_3.csv", rows_35_3)


# ===========================================================================
# Phase 3 feedback CSVs
# These are read by phase_3_processor.py using safe_open_csv (CSV only).
# Row[0]: "Reported Name & DOB" (IDs file)  or "Reported ID" (Names file)
# Row[1]: "Reported IDs"       (IDs file)  or "Reported Names & DOBs" (Names)
# Row[4]: Incident codes (pipe-delimited)
# Row[6]: Client Confirmed Correction  ← written by processor
# Row[7]: Client Confirmed Correction Fields ← written by processor
# ===========================================================================

_P3_IDS_HEADER = [
    "Reported Name & DOB",
    "Reported IDs",
    "Suggested Correction (Best Efforts)",
    "Incident Types",
    "Incident Codes",
    "Totals",
    "Client Confirmed Correction",
    "Client Confirmed Correction Fields",
    "Client Comments",
]

_P3_NAMES_HEADER = [
    "Reported ID",
    "Reported Names & DOBs",
    "Suggested Correction (Best Efforts)",
    "Incident Types",
    "Incident Codes",
    "Totals",
    "Client Confirmed Correction",
    "Client Confirmed Correction Fields",
    "Client Comments",
]


def create_phase_3_feedback(base_dir: Path) -> None:
    out_dir = base_dir / "phase_3_feedback"
    _make_dir(out_dir)

    # ── IDs Summary (phase_3_processor input) ────────────────────────────────
    # Row[0] = "FirstName~Surname~YYYY-MM-DD"
    # Row[1] = "TYPE:ID" or newline-delimited "TYPE:ID\nTYPE:ID"
    ids_rows = [_P3_IDS_HEADER] + [
        # 1. Single NIDN — no correction needed
        ["ALICE~JOHNSON~1980-04-15", "NIDN:GBAJ123456B",
         None, "Invalid FN", "13_1", "3", "No Change", None, None],

        # 2. Two IDs for same person — CONCAT is non-applicable; correction is the NIDN
        ["ALICE~JOHNSON~1980-04-15",
         "NIDN:GBAJ123456B\nCONCAT:GB19800415ALICJOHNSO",
         None, "Non-applicable ID type", "7_36", "1",
         "GBAJ123456B", "ID", None],

        # 3. CONCAT only — non-applicable + below min DOB; no correction
        ["BOB~SMITH~1975-06-22", "CONCAT:GB19750622BOBSMITH",
         None, "Non-applicable ID type|Below min DOB", "7_36|11_2", "2",
         "No Change", None, None],

        # 4. Inconsistent ID — correction provided
        ["CHARLIE~BROWN~1965-11-30", "NIDN:GBCB987654C",
         None, "Inconsistent ID", "7_66", "5",
         "GBCB987654C", "ID", None],

        # 5. LEI buyer — corporate entity; no person ID correction
        ["N/A~N/A~N/A", "LEI:529900T8BM49AURSDO55",
         None, "Invalid ID format", "7_3", "1", "No Change", None, "Cannot verify"],

        # 6. NIDN — no correction yet (blank col 6/7) — processor will fill
        ["EDWARD~JONES~1990-02-14", "NIDN:GBEJ654321D",
         None, "Missing first name", "13_1", "2", None, None, None],

        # 7. Multiple incident codes; correction present
        ["FIONA~APPLE~1977-08-03", "NIDN:GBFA111222A",
         None, "Invalid FN|Inconsistent ID", "13_1|7_66", "8",
         "GBFA111222A", "ID", None],

        # 8. Single NIDN — no correction (non-reportable)
        ["GEORGE~MARTIN~1943-06-03", "NIDN:GBGM333444B",
         None, "Non-reportable", "16_18", "1", "No Change", None, None],

        # 9. CONCAT with DOB embedded — format correction
        ["HARRIET~WILSON~1988-07-19",
         "CONCAT:GB19880719HARRIWILS",
         None, "Non-applicable ID type", "7_36", "3",
         "GBHW567890C", "ID", None],

        # 10. Buyer DM code — correction on decision maker
        ["IAN~TAYLOR~1955-03-25", "NIDN:GBIT345678D",
         None, "Invalid DM code", "12_17", "1",
         "GBIT345678D", "Buyer decision maker code", None],

        # 11. Person not found in incident template — blank correction (unmatched)
        ["JULIA~ROBERTS~1967-10-28", "NIDN:GBJR456789E",
         None, "Inconsistent ID", "7_66", "2", None, None, None],

        # 12. Swedish passport — non-UK ID
        ["KARL~ANDERSSON~1972-04-11", "PASS:SE72041199999",
         None, "Non-applicable ID type", "7_36", "1",
         "GBKA789012F", "ID", None],
    ]
    _write_csv(
        out_dir / "Replay_2025Q3_PHASE 3_Inconsistent_IDs_Summary_FINAL.csv",
        ids_rows,
    )

    # ── Names Summary (phase_3_processor input) ───────────────────────────────
    # Row[0] = "ID_value~ID_type"
    # Row[1] = "First:Surname:YYYY-MM-DD"  (newline-delimited for multiple)
    names_rows = [_P3_NAMES_HEADER] + [
        # 1. CONCAT ID — non-applicable type; no correction
        ["AF19650501JAMES#HARRI~CONCAT",
         "JAMES,WILLIAM:HARRISON:1965-05-01",
         None, "Non-applicable ID type", "16_18", "1", "No Change", None, None],

        # 2. CONCAT — correction to NIDN
        ["AF19750622BOB#SMITH~CONCAT",
         "BOB:SMITH:1975-06-22",
         None, "Non-applicable ID type", "16_18", "3",
         "GBBS123456A", "ID", None],

        # 3. INTC — internal counterparty; no lookup needed
        ["INTC~",
         "INTC:INTC:",
         None, "N/A", "7_11", "10", "No Change", None, "Internal to firm"],

        # 4. NIDN — multiple names for same ID (name inconsistency)
        ["GBCB987654C~NIDN",
         "CHARLIE:BROWN:1965-11-30\nCHARLES:BROWN:1965-11-30",
         None, "Inconsistent name", "13_1", "5",
         "CHARLIE", "FN", None],

        # 5. LEI — corporate, no person correction
        ["529900T8BM49AURSDO55~LEI",
         "N/A:N/A:",
         None, "Invalid ID format", "7_3", "2", "No Change", None, None],

        # 6. NIDN — first name correction
        ["GBAJ123456B~NIDN",
         "ALICE:JOHNSON:1980-04-15",
         None, "Non-applicable ID type", "7_36", "3",
         "No Change", None, None],

        # 7. CONCAT — multiple names DOBs; correction needed
        ["GB19880719HARRIWILS~CONCAT",
         "HARRIET:WILSON:1988-07-19\nHARRIET:WILSONE:1988-07-19",
         None, "Inconsistent name", "13_1", "2",
         "HARRIET", "FN", None],

        # 8. NIDN — not found in incident template (blank correction)
        ["GBEJ654321D~NIDN",
         "EDWARD:JONES:1990-02-14",
         None, "Missing first name", "13_1", "2", None, None, None],

        # 9. CONCAT — surname correction
        ["AF19720411KARL#ANDE~CONCAT",
         "KARL:ANDERSSON:1972-04-11\nKARL:ANDERSEN:1972-04-11",
         None, "Inconsistent surname", "13_1", "3",
         "ANDERSSON", "SN", None],

        # 10. NIDN with pipe-delimited incident codes
        ["GBFA111222A~NIDN",
         "FIONA:APPLE:1977-08-03",
         None, "Invalid FN|Inconsistent ID", "13_1|7_66", "5",
         "FIONA", "FN", None],
    ]
    _write_csv(
        out_dir / "Replay_2025Q3_PHASE 3_Inconsistent_Names_Summary_FINAL.csv",
        names_rows,
    )


# ===========================================================================
# Phase 3 Final Lookup input files
# These are the PROCESSED outputs from phase_3_processor (corrections filled in)
# PLUS the UnaVista CSV (87 columns).
# ===========================================================================

def create_phase_3_final_lookup(base_dir: Path) -> None:
    out_dir = base_dir / "phase_3_final_lookup"
    _make_dir(out_dir)

    # ── Processed IDs Summary (phase_3_final_lookup input) ───────────────────
    # Same format as phase_3_feedback IDs Summary but corrections are already filled.
    # The lookup verifies these corrections against the UnaVista data.
    processed_ids = [_P3_IDS_HEADER] + [
        # 1. PASS — Alice Johnson, correction GBAJ123456B matches UnaVista Buyer ID
        ["ALICE~JOHNSON~1980-04-15", "NIDN:GBAJ123456B",
         None, "Invalid FN", "13_1", "3", "GBAJ123456B", "ID", None],

        # 2. FAIL — Bob Smith, correction does not match UnaVista (wrong correction)
        ["BOB~SMITH~1975-06-22", "NIDN:GBBS654321A",
         None, "Inconsistent ID", "7_66", "2",
         "GBNEWWRONG99", "ID", None],

        # 3. No change — Charlie Brown
        ["CHARLIE~BROWN~1965-11-30", "NIDN:GBCB987654C",
         None, "Inconsistent ID", "7_66", "5", "No Change", None, None],

        # 4. PASS — Fiona Apple
        ["FIONA~APPLE~1977-08-03", "NIDN:GBFA111222A",
         None, "Invalid FN|Inconsistent ID", "13_1|7_66", "8",
         "GBFA111222A", "ID", None],

        # 5. Client not found — Diana Wells (no UnaVista row matches this ID)
        ["DIANA~WELLS~1988-12-05", "NIDN:GBDW111222B",
         None, "Inconsistent ID", "7_66", "1", "GBDW111222B", "ID", None],

        # 6. PASS — Harriet Wilson, ID corrected
        ["HARRIET~WILSON~1988-07-19", "CONCAT:GB19880719HARRIWILS",
         None, "Non-applicable ID type", "7_36", "3",
         "GBHW567890C", "ID", None],

        # 7. No change — George Martin (non-reportable)
        ["GEORGE~MARTIN~1943-06-03", "NIDN:GBGM333444B",
         None, "Non-reportable", "16_18", "1", "No Change", None, None],

        # 8. PASS — Edward Jones, DM correction
        ["EDWARD~JONES~1990-02-14", "NIDN:GBEJ654321D",
         None, "Invalid DM code", "12_17", "2",
         "GBEJ654321D", "Buyer decision maker code", None],
    ]
    _write_csv(
        out_dir / "Replay_2025Q3_Inconsistent_IDs_Summary_FINAL.csv",
        processed_ids,
    )

    # ── Processed Names Summary (phase_3_final_lookup input) ─────────────────
    processed_names = [_P3_NAMES_HEADER] + [
        # 1. PASS — Bob Smith ID correction matches UnaVista seller
        ["AF19750622BOB#SMITH~CONCAT",
         "BOB:SMITH:1975-06-22",
         None, "Non-applicable ID type", "16_18", "3",
         "GBBS123456A", "ID", None],

        # 2. PASS — Alice Johnson first name in UnaVista matches No Change
        ["GBAJ123456B~NIDN",
         "ALICE:JOHNSON:1980-04-15",
         None, "Non-applicable ID type", "7_36", "3",
         "No Change", None, None],

        # 3. FAIL — Harriet Wilson surname correction wrong
        ["GB19880719HARRIWILS~CONCAT",
         "HARRIET:WILSON:1988-07-19\nHARRIET:WILSONE:1988-07-19",
         None, "Inconsistent name", "13_1", "2",
         "WRONGSURNAME", "SN", None],

        # 4. PASS — Charlie Brown first name correction
        ["GBCB987654C~NIDN",
         "CHARLIE:BROWN:1965-11-30\nCHARLES:BROWN:1965-11-30",
         None, "Inconsistent name", "13_1", "5",
         "CHARLIE", "FN", None],

        # 5. No change — INTC
        ["INTC~",
         "INTC:INTC:",
         None, "N/A", "7_11", "10", "No Change", None, "Internal to firm"],

        # 6. Client not found — unknown ID
        ["GBUNK999888A~NIDN",
         "UNKNOWN:PERSON:1960-01-01",
         None, "Invalid FN", "13_1", "1",
         "UNKNOWN", "FN", None],

        # 7. PASS — Fiona Apple FN correction
        ["GBFA111222A~NIDN",
         "FIONA:APPLE:1977-08-03",
         None, "Invalid FN|Inconsistent ID", "13_1|7_66", "5",
         "FIONA", "FN", None],

        # 8. PASS — Karl Andersson surname correction
        ["AF19720411KARL#ANDE~CONCAT",
         "KARL:ANDERSSON:1972-04-11\nKARL:ANDERSEN:1972-04-11",
         None, "Inconsistent surname", "13_1", "3",
         "ANDERSSON", "SN", None],
    ]
    _write_csv(
        out_dir / "Replay_2025Q3_Inconsistent_Names_Summary_FINAL.csv",
        processed_names,
    )

    # ── UnaVista CSV (87 columns) ─────────────────────────────────────────────
    # Col 1 = Transaction Reference Number
    # Col 8 = Buyer ID  (lookup key for IDs summary)
    # Col 10 = Buyer First Name, 11 = Buyer Surname, 12 = Buyer DOB
    # Col 21 = Seller ID  (lookup key for Names summary)
    # Col 23 = Seller First Name, 24 = Seller Surname, 25 = Seller DOB
    uv_header = [
        "Report Status",            # 0
        "Transaction Reference Number",  # 1
        "Venue Transaction ID",     # 2
        "Submitting Entity ID",     # 3
        "Executing Entity ID",      # 4
        "Investment Firm Indicator",  # 5
        "Buyer ID Type",            # 6
        "Buyer ID Sub Type",        # 7
        "Buyer ID",                 # 8
        "Buyer Country of Branch",  # 9
        "Buyer First Name",         # 10
        "Buyer Surname",            # 11
        "Buyer DOB",                # 12
        "Buyer Decision Maker ID Type",      # 13
        "Buyer Decision Maker ID Sub Type",  # 14
        "Buyer Decision Maker ID",           # 15
        "Buyer Decision Maker First Name",   # 16
        "Buyer Decision Maker Surname",      # 17
        "Buyer Decision Maker DOB",          # 18
        "Seller ID Type",           # 19
        "Seller ID Sub Type",       # 20
        "Seller ID",                # 21
        "Seller Country of Branch", # 22
        "Seller First Name",        # 23
        "Seller Surname",           # 24
        "Seller DOB",               # 25
        "Seller Decision Maker ID Type",     # 26
        "Seller Decision Maker ID Sub Type", # 27
        "Seller Decision Maker ID",          # 28
        "Seller Decision Maker First Name",  # 29
        "Seller Decision Maker Surname",     # 30
        "Seller Decision Maker DOB",         # 31
        "Order Transmission Indicator",      # 32
        "Buyer Transmitter ID",     # 33
        "Seller Transmitter ID",    # 34
        "Trading Date Time",        # 35
        "Trading Capacity",         # 36
        "Quantity",                 # 37
        "Quantity Type",            # 38
        "Quantity Currency",        # 39
        "Derivative Notional Change",  # 40
        "Price",                    # 41
        "Price Type",               # 42
        "Price Currency",           # 43
        "Net Amount",               # 44
        "Venue",                    # 45
        "Country of Branch",        # 46
        "Up-Front Payment",         # 47
        "Up-Front Payment Currency", # 48
        "Complex Trade Component ID",  # 49
        "Instrument ID Type",       # 50
        "Instrument ID",            # 51
        "Instrument Name",          # 52
        "Instrument Classification",  # 53
        "Notional Currency 1",      # 54
        "Notional Currency 2",      # 55
        "Notional Currency 2 Type", # 56
        "Price Multiplier",         # 57
        "UV Instrument Classification",  # 58
        "Underlying Instrument ID", # 59
        "UV Index Classification",  # 60
        "Underlying Index ID",      # 61
        "Underlying Index Name",    # 62
        "Underlying Index Term",    # 63
        "Option Type",              # 64
        "Strike Price",             # 65
        "Strike Price Type",        # 66
        "Strike Price Currency",    # 67
        "Option Style",             # 68
        "Maturity Date",            # 69
        "Expiry Date",              # 70
        "Delivery Type",            # 71
        "Investment Decision ID Type",      # 72
        "Investment Decision ID Sub Type",  # 73
        "Investment Decision ID",           # 74
        "Investment Decision Country of Branch",  # 75
        "Firm Execution ID Type",   # 76
        "Firm Execution ID Sub Type",  # 77
        "Firm Execution ID",        # 78
        "Firm Execution Country of Branch",  # 79
        "Waiver Indicator",         # 80
        "Short Selling Indicator",  # 81
        "OTC Post Trade Indicator", # 82
        "Commodity Derivative Indicator",  # 83
        "SFT Indicator",            # 84
        "Internal Client Identification",  # 85
        "Data Category",            # 86
    ]
    assert len(uv_header) == 87, f"UnaVista header must be 87 cols, got {len(uv_header)}"

    SUBMITTER = "213800SAMPLE0001LEI1"
    INSTRUMENT = "GB0007980591"

    def uv_row(
        status, txn_ref,
        buyer_id_type, buyer_id_sub, buyer_id,
        buyer_country, buyer_fn, buyer_sn, buyer_dob,
        buyer_dm_id_type="", buyer_dm_id_sub="", buyer_dm_id="",
        buyer_dm_fn="", buyer_dm_sn="", buyer_dm_dob="",
        seller_id_type="I", seller_id_sub="", seller_id="INTC",
        seller_country="", seller_fn="", seller_sn="", seller_dob="",
        seller_dm_id_type="", seller_dm_id_sub="", seller_dm_id="",
        seller_dm_fn="", seller_dm_sn="", seller_dm_dob="",
        trading_cap="AOTC", quantity=100, price=10.0, net_amount=1000.0,
        instrument=INSTRUMENT,
    ):
        row = [""] * 87
        row[0] = status
        row[1] = txn_ref
        row[2] = ""
        row[3] = SUBMITTER
        row[4] = SUBMITTER
        row[5] = "True"
        row[6] = buyer_id_type
        row[7] = buyer_id_sub
        row[8] = buyer_id
        row[9] = buyer_country
        row[10] = buyer_fn
        row[11] = buyer_sn
        row[12] = buyer_dob
        row[13] = buyer_dm_id_type
        row[14] = buyer_dm_id_sub
        row[15] = buyer_dm_id
        row[16] = buyer_dm_fn
        row[17] = buyer_dm_sn
        row[18] = buyer_dm_dob
        row[19] = seller_id_type
        row[20] = seller_id_sub
        row[21] = seller_id
        row[22] = seller_country
        row[23] = seller_fn
        row[24] = seller_sn
        row[25] = seller_dob
        row[26] = seller_dm_id_type
        row[27] = seller_dm_id_sub
        row[28] = seller_dm_id
        row[29] = seller_dm_fn
        row[30] = seller_dm_sn
        row[31] = seller_dm_dob
        row[32] = "False"
        row[35] = "2025-07-01T08:00:00.000000Z"
        row[36] = trading_cap
        row[37] = str(quantity)
        row[38] = "Unit"
        row[41] = str(price)
        row[42] = "MntryValAmt"
        row[43] = "GBP"
        row[44] = str(net_amount)
        row[45] = "XOFF"
        row[50] = "FinInstrm.Id"
        row[51] = instrument
        row[86] = "N"
        return row

    uv_data = [uv_header]

    # 1. PASS — Alice Johnson, Buyer NIDN matches IDs Summary → correction matches UnaVista
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_001",
        "N", "NIDN", "GBAJ123456B", "GB", "ALICE", "JOHNSON", "1980-04-15",
        quantity=175, price=3.642, net_amount=637.35,
    ))

    # 2. FAIL — Bob Smith, Buyer NIDN is in Names Summary; correction GBNEWWRONG99 ≠ UnaVista
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_002",
        "N", "NIDN", "GBBS654321A", "GB", "BOB", "SMITH", "1975-06-22",
        quantity=50, price=20.0, net_amount=1000.0,
    ))

    # 3. No change — Charlie Brown; UnaVista has the same ID → "No change" result
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_003",
        "N", "NIDN", "GBCB987654C", "GB", "CHARLIE", "BROWN", "1965-11-30",
        seller_id_type="I", seller_id="INTC",
        quantity=200, price=5.0, net_amount=1000.0,
    ))

    # 4. PASS — Fiona Apple, Buyer NIDN matches IDs Summary → correction matches
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_004",
        "N", "NIDN", "GBFA111222A", "GB", "FIONA", "APPLE", "1977-08-03",
        quantity=80, price=12.5, net_amount=1000.0,
    ))

    # 5. Client not found — Diana Wells; no IDs Summary entry with GBDW111222B
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_005",
        "N", "NIDN", "GBDW111222B", "GB", "DIANA", "WELLS", "1988-12-05",
        quantity=30, price=33.33, net_amount=999.9,
    ))

    # 6. PASS — Harriet Wilson; buyer CONCAT ID in IDs Summary; correction GBHW567890C
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_006",
        "N", "NIDN", "GBHW567890C", "GB", "HARRIET", "WILSON", "1988-07-19",
        quantity=40, price=25.0, net_amount=1000.0,
    ))

    # 7. PASS — Edward Jones, buyer DM correction matches UnaVista DM ID
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_007",
        "N", "NIDN", "GBEJ654321D", "GB", "EDWARD", "JONES", "1990-02-14",
        buyer_dm_id_type="N", buyer_dm_id_sub="NIDN", buyer_dm_id="GBEJ654321D",
        buyer_dm_fn="EDWARD", buyer_dm_sn="JONES", buyer_dm_dob="1990-02-14",
        quantity=120, price=8.33, net_amount=999.6,
    ))

    # 8. No change — George Martin; no correction in IDs Summary
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_008",
        "N", "NIDN", "GBGM333444B", "GB", "GEORGE", "MARTIN", "1943-06-03",
        quantity=300, price=3.33, net_amount=999.0,
    ))

    # 9. PASS — Bob Smith as SELLER (from Names Summary)
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_009",
        "L", "", "213800SAMPLE0001LEI1", "", "", "", "",
        seller_id_type="N", seller_id_sub="NIDN",
        seller_id="GBBS123456A", seller_country="GB",
        seller_fn="BOB", seller_sn="SMITH", seller_dob="1975-06-22",
        quantity=90, price=11.1, net_amount=999.0,
    ))

    # 10. FAIL — Harriet Wilson as SELLER (wrong surname correction in Names Summary)
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_010",
        "L", "", "213800SAMPLE0001LEI1", "", "", "", "",
        seller_id_type="N", seller_id_sub="NIDN",
        seller_id="GBHW567890C", seller_country="GB",
        seller_fn="HARRIET", seller_sn="WILSON", seller_dob="1988-07-19",
        quantity=55, price=18.18, net_amount=999.9,
    ))

    # 11. PASS — Charlie Brown as SELLER (FN correction "CHARLIE")
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_011",
        "L", "", "213800SAMPLE0001LEI1", "", "", "", "",
        seller_id_type="N", seller_id_sub="NIDN",
        seller_id="GBCB987654C", seller_country="GB",
        seller_fn="CHARLIE", seller_sn="BROWN", seller_dob="1965-11-30",
        quantity=250, price=4.0, net_amount=1000.0,
    ))

    # 12. PASS — Fiona Apple as SELLER (FN correction "FIONA")
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_012",
        "L", "", "213800SAMPLE0001LEI1", "", "", "", "",
        seller_id_type="N", seller_id_sub="NIDN",
        seller_id="GBFA111222A", seller_country="GB",
        seller_fn="FIONA", seller_sn="APPLE", seller_dob="1977-08-03",
        quantity=65, price=15.38, net_amount=999.7,
    ))

    # 13. PASS — Karl Andersson (ANDERSSON surname correction)
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_013",
        "N", "NIDN", "GBKA789012F", "GB", "KARL", "ANDERSSON", "1972-04-11",
        seller_id_type="I", seller_id="INTC",
        quantity=150, price=6.67, net_amount=1000.5,
    ))

    # 14. CANC — cancellation row; buyer/seller data sparse
    uv_data.append(uv_row(
        "CANC", "UVISTA_P3_014",
        "N", "NIDN", "GBAJ123456B", "GB", "ALICE", "JOHNSON", "1980-04-15",
        quantity=175, price=3.642, net_amount=637.35,
    ))

    # 15. LEI buyer — corporate entity (no person ID lookup)
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_015",
        "L", "", "529900T8BM49AURSDO55", "", "", "", "",
        seller_id_type="I", seller_id="INTC",
        quantity=1000, price=1.0, net_amount=1000.0,
    ))

    # 16. PASS — Alice Johnson (second transaction, same buyer)
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_016",
        "N", "NIDN", "GBAJ123456B", "GB", "ALICE", "JOHNSON", "1980-04-15",
        quantity=50, price=20.0, net_amount=1000.0,
        instrument="GB0002634946",
    ))

    # 17. Unrelated row — no IDs/Names summary entry; not tested
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_017",
        "N", "NIDN", "GBUNRELATED1A", "GB", "UNRELATED", "PERSON", "1960-06-15",
        quantity=10, price=100.0, net_amount=1000.0,
    ))

    # 18. NEWT with INTC seller — common AOTC pattern
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_018",
        "N", "NIDN", "GBGM333444B", "GB", "GEORGE", "MARTIN", "1943-06-03",
        seller_id_type="I", seller_id="INTC",
        quantity=400, price=2.5, net_amount=1000.0,
        instrument="GB0033537902",
    ))

    # 19. MODI — modification; update from original
    uv_data.append(uv_row(
        "MODI", "UVISTA_P3_019",
        "N", "NIDN", "GBBS654321A", "GB", "BOB", "SMITH", "1975-06-22",
        quantity=50, price=20.0, net_amount=1000.0,
    ))

    # 20. PASS — Bob Smith buyer (from IDs summary — correction GBNEWWRONG99 makes this FAIL)
    #     Intentionally uses original ID so the IDs summary correction will mismatch
    uv_data.append(uv_row(
        "NEWT", "UVISTA_P3_020",
        "N", "NIDN", "GBBS654321A", "GB", "BOB", "SMITH", "1975-06-22",
        quantity=70, price=14.28, net_amount=999.6,
        instrument="GB0005405286",
    ))

    _write_csv(
        out_dir / "UnaVista_MiFIR_Manual_Corrections_423_20180406111252.(264).csv",
        uv_data,
    )


# ===========================================================================
# Main
# ===========================================================================

def main() -> None:
    print("Generating replay sample data...")
    print(f"Output root: {BASE_DIR}")
    print()

    print("=== Phase 2 feedback XLSX ===")
    create_phase_2_feedback(BASE_DIR)

    print()
    print("=== Phase 2 incident template CSVs ===")
    create_phase_2_incident_templates(BASE_DIR)

    print()
    print("=== Phase 3 feedback CSVs ===")
    create_phase_3_feedback(BASE_DIR)

    print()
    print("=== Phase 3 final lookup files ===")
    create_phase_3_final_lookup(BASE_DIR)

    print()
    print("Done.  All replay sample files created successfully.")
    print()
    print("Directories created:")
    for sub in ("phase_2_feedback", "phase_2_incident_templates",
                "phase_3_feedback", "phase_3_final_lookup"):
        d = BASE_DIR / sub
        files = sorted(d.iterdir())
        print(f"  {sub}/  ({len(files)} files)")
        for f in files:
            print(f"    {f.name}")


if __name__ == "__main__":
    main()
